#!/usr/bin/env python
# coding: utf-8

# In[ ]:


get_ipython().run_line_magic('matplotlib', 'qt')
#%matplotlib inline
import os
import csv
import numpy as np
import re
import datetime
import fnmatch
from pathlib import Path
from itertools import groupby
import copy
import itertools
import distance
import nltk
#nltk.download('stopwords')
import pandas as pd
import matplotlib.pyplot as plt
import pickle

from nltk.stem import SnowballStemmer
from nltk.stem import snowball

from itertools import *
from operator import *

from openpyxl import load_workbook


# # Dictionaries

# In[ ]:


# exceptional removal of particular extra sentences not typed by the user 
dict_phraseStim = {
    #'2019-02-05-14-10-39_2ndPart_2' : [1, 2, 3, 4, 5, 6, 9, 10],
    #'2019-01-14-14-58-30' : [0], # ys, session_trial ()
    '2019-01-16-16-36-17_1stPart_2' : [-1], # af_session1
    '2019-01-16-17-00-12_2ndPart_2': [1], # af_session1
    '2019-01-17-15-27-20_1stPart_2' : [4], # Af session2
    '2019-01-17-16-03-27_2ndPart_2' : [0, 1, 2], # Af session2
    '2019-02-06-11-25-41_1' : [7],               # aq_session1    
    '2019-02-08-11-33-53_1stPart_1' : [1],  # aq session3_1_part1
    '2019-02-08-12-11-34_2ndPart_1' : [0, 1, 2, 3],  # aq session3_1_part2
    '2019-01-31-09-37-5_2ndPart_2' : range(1,5), # bh1, session 4 , all sentences except the first one deleted
    '2019-01-31-09-22-49_1stPart_2' : [4],  # bh1_session4_2_part1
    '2019-02-21-16-09-44_1stPart_1' : [1], # bh2_session1
    '2019-02-21-16-22-22_2ndPart_1' : [2, 3, 4],# bh2_session1
    '2019-02-28-17-03-53_1stPart_2' : [2],       # bh2_session3
    '2019-02-28-17-24-2_2ndPart_2' : [0, 2],     # bh2_session3
    '2019-02-14-13-28-20_1stPart_2' : [2], # cw_session3_2_part1
    '2019-02-14-13-57-41_2ndPart_2' : [0, 2, 3], # cw_session3_2_part2
    '2019-02-21-15-01-4_1stPart_1' : [0],        # le_session3
    '2019-02-21-15-25-56_2ndPart_1' : [1],        # le_session3
    '2019-02-18-10-28-35_2' : [0],               # ls2_session4 # picture not described
    '2019-02-05-14-00-27_1stPart_2' : [3],        # mh_session1
    '2019-02-05-14-10-39_2ndPart_2' : [0, 1, 3],   # mh_session1
    '2019-02-08-10-51-3_1stPart_1' : [4],        # mn_session1
    '2019-02-08-11-05-7_2ndPart_1' : [0, 2, 3, 4], # mn_session1
    '2019-02-19-10-34-7_1stPart_1' : [3],          # mn_session3
    '2019-02-19-10-56-43_2ndPart_1' : [1, 2, 3, 4], # mn_session3
    '2019-01-16-15-18-0_1' : [4],            # no_session1
    '2019-02-19-17-10-45_1' : [3],                  # ph_session5
    '2019-01-29-13-25-4_1' : [3],        # ph_session2
    '2019-03-07-16-44-5_2' : [1],                   # rh_session1
    '2019-03-14-13-56-56_2' : [2],                  # rh_session3
    '2019-01-14-15-07-21_1' : [4], # ys_session1
    '2019-01-16-15-18-50_1stPart_1' : [3, 4], # ys_session2
    '2019-01-16-15-42-51_2ndPart_1' : [2], # ys_session2
    '2019-01-30-11-22-25_1' : [3, 5, 7],          # ys_session4
    '2019-01-30-11-22-25_1' : [4, 6, 7] # ys, session 4
}

# exceptional removal of sentences/words typed by the user, but then deleted everything to have a blank scratchpad

dict_phraseUser = {
    "2019-02-06-15-44-15_1" : [2, 3, 6], 
    "2019-02-06-16-19-9_2" : [1, 3, 6, 7],
    "2019-02-12-11-21-21_2" : [0],
    "2019-02-14-14-28-49_1" : [0, 2, 3], # ac_session3_1
    "2019-02-14-14-45-49_2" : [0, 5, 6], # ac_session3_2
    '2019-01-29-14-19-26_1' : [0, 3, 4], # bh1_session2_1
    '2019-01-29-14-40-36_2' : [0, 1, 2], # bh1_session2_2
    '2019-01-30-14-29-29_2' : [4],       # bh1_session3_2
    '2019-01-31-09-12-2_1' : [3],         # bh1_session4_1
    '2019-01-31-09-22-49_1stPart_2' : [4], # bh1_session4_2_part1
    '2019-03-05-09-15-11_1' : [1],         # bh2_session5_1
    '2019-03-05-09-15-11_2' : [1],        # bh2_session5_2
    '2019-02-21-15-55-56_2' : [2],       # ch_session5_2
    '2019-01-30-15-19-36_2' : [1],       # jm_session2_1
    '2019-01-30-15-04-30_1' : [0],         # jm_session2_2
    '2019-01-16-15-18-50_1stPart_1' : [1],  # ys_session2
    '2019-01-16-15-42-51_2ndPart_1' : [0], # ys_session2
    '2019-01-30-11-22-25_1' : [2, 4],       # ys_session4
    '2019-01-30-11-57-3_2' : [0] ,          # ys_session4
    '2019-01-31-13-13-2_1' : [4],           # ys_session5
    '2019-01-30-10-20-32_1' : [0, 1, 2, 3, 4, 5], # no_session4
    '2019-01-30-10-46-38_2' : [0],          # 
    '2019-02-28-17-03-53_1stPart_2' : [2],   # bh2_session3
    '2019-03-12-09-30-5_1' : [0],            # kj_session3
    '2019-02-13-15-20-38_1' : [0, 1, 2, 3, 6], # ls1_session3
    '2019-02-18-10-25-52_1' : [1],              # ls2_session4
    '2019-02-18-10-46-26_2' : [0],            # ls2_session4
    '2019-01-29-13-25-4_1' : [0, 1, 7],        # ph_session2
    '2019-01-29-13-43-50_2' : [0],              # ph_session2
    '2019-03-07-16-17-30_1' : [0],              # rh_session1
    '2019-03-07-16-44-5_2' : [0, 1],         # rh_session1
    '2019-03-14-13-56-56_2' : [0, 1, 3]         # rh_session3
}

# key selection can have extra selections of NextPhrase at the end
dict_keySelectionOfNextPhrase = {
    "2019-02-11-11-18-30_1" : [12, 13], # ac_session1
    "2019-01-16-17-00-12_2ndPart_2" : [12], # af_session1
    "2019-01-17-15-27-20_1stPart_2" : [12], # af_session2
    "2019-02-06-16-19-9_2" : [12], # af_session3
    "2019-02-12-11-07-43_1" : [12], # af_session4
    "2019-02-27-15-08-32_1" : [12], # af_session5
    "2019-01-28-14-30-44_1" : [12], # bh1_session1
    "2019-02-21-16-22-22_2ndPart_1" : [12], # bh2_session1
    "2019-02-18-14-02-56_2" : [12], # le_session1
    "2019-02-19-10-03-14_1" : [12], # le_session2
    "2019-02-08-11-05-7_2ndPart_1" : [12], # mn_session1
    "2019-02-08-11-12-51_2" : [12, 13], # mn_session1
    "2019-02-15-11-38-22_1" : [12, 13], # mn_session2
    "2019-02-15-11-54-25_2" : [12], # mn_session2
    "2019-01-16-15-18-0_1" : [12], # no_session1
    "2019-01-28-13-31-51_1" : [12], # ph_session1
    "2019-01-28-13-49-14_2" : [12], # ph_session1
    "2019-01-14-15-07-21_1" : [12], # ys_session1
    "2019-01-17-15-05-1_1" : [12], # ys_session3
    "2019-01-30-11-22-25_1" : [12], # ys_session4
    "2019-01-31-13-32-2_2" : [12], # ys_session5
}


# key selection when participants skips some sentences
dict_keySelectionNotCompleted = {
    "2019-01-16-16-36-17_1stPart_2" : [0, 1, 3, 5, 7], # af_session1 ---- last sentence is not finished
    "2019-01-16-17-00-12_2ndPart_2" : [0, 1, 3, 4, 5, 7, 9, 11], # af_session1
    "2019-01-17-15-27-20_1stPart_2" : [0, 1, 3, 5, 7, 9, 11], # af_session2 
    "2019-01-17-16-03-27_2ndPart_2" : [0, 1, 2, 3, 4, 5, 6, 7, 9, 11], # af_session2
    "2019-02-08-11-33-53_1stPart_1" : [0, 1, 3, 4, 5, 7, 9, 11], # aq_session3
    "2019-02-08-12-11-34_2ndPart_1" : [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 11], # aq_session3
    "2019-01-28-14-30-44_1" : [0, 1, 3, 5], # bh1_session1
    "2019-01-31-09-22-49_1stPart_2": [0, 1, 3, 5, 7, 9, 10, 11], # bh1_session4
    "2019-01-31-09-37-5_2ndPart_2" : [0, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11], # bh1_session4
    "2019-02-21-16-09-44_1stPart_1" : [0, 1, 3, 4, 5, 7, 9, 11], # bh2_session1
    "2019-02-21-16-22-22_2ndPart_1" : [0, 1, 3, 5, 6, 7, 8, 9, 10, 11], # bh2_session1
    "2019-02-28-17-03-53_1stPart_2" : [0, 1, 3, 5, 6, 7, 9, 11], # bh2_session3
    "2019-02-28-17-24-2_2ndPart_2" : [0, 1, 2, 3, 5], # bh2_session3     ----
    "2019-02-14-13-28-20_1stPart_2" : [0, 1, 3, 5, 6, 7, 9, 11], # cw_session3
    "2019-02-14-13-57-41_2ndPart_2" : [0, 1, 2, 3, 5, 6, 7, 8, 9, 11], # cw_session3
    "2019-02-21-15-01-4_1stPart_1" : [0, 1, 2, 3, 5, 7, 9, 11], # le_session3
    "2019-02-21-15-25-56_2ndPart_1" : [0, 1, 3], # le_session3       ----
    "2019-02-05-14-00-27_1stPart_2" : [0, 1, 3, 5, 7, 8], # mh_session1
    "2019-02-05-14-10-39_2ndPart_2" : [0, 1, 2, 3, 4, 5, 7, 8, 9, 11], # mh_session1
    "2019-02-08-10-51-3_1stPart_1" : [0, 1, 3, 5, 7, 9, 10, 11], # mn_session1
    "2019-02-08-11-05-7_2ndPart_1" : [0, 1, 2, 3, 5, 6, 7, 8, 9, 10, 11], # mn_session1
    "2019-02-19-10-34-7_1stPart_1" : [0, 1, 3, 5, 7, 8, 9, 11], # mn_session3
    "2019-02-19-10-56-43_2ndPart_1" : [0, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11], # mn_session3
    "2019-01-29-13-25-4_1" : [0, 1, 3, 5, 7], # ph_session2  -- sessions where there are less score questions 
    # and more sentences typed
    "2019-01-16-15-18-50_1stPart_1" : [0, 1, 3, 5, 7, 8, 9, 10], # ys_session2
    "2019-01-17-15-05-1_1" : [0, 1, 3, 5],  # ys_session3  -- sessions where there are less score questions 
    # and more sentences typed
    "2019-02-06-11-25-41_1" : [0, 1, 3, 5, 11], # aq_session1 -- sessions where there are less score questions 
    # and more sentences typed
    "2019-01-16-15-42-51_2ndPart_1" : [0, 1, 2, 5], # ys_session2 -- different for reading and writing, this one is for
    # writing
    '2019-01-30-11-22-25_1' : [0, 1, 3, 5, 7, 9, 11]   # ys_session4 -- sessions where there are less score questions 
    # and more sentences typed
   
}

# dictionary for phrase removal just like in the dict_phraseStim, but since not all participants require that, some that 
# do, are added to this new dictionary here
dict_keySelection_phraseStim = {
    '2019-01-17-15-27-20_1stPart_2' : [4], # Af session2
    '2019-01-16-15-18-0_1' : [4],        # no_session1
    '2019-02-19-17-10-45_1' : [3],                  # ph_session5
    '2019-03-07-16-44-5_2' : [1],        # rh_session1
    '2019-03-14-13-56-56_2' : [2],              # rh_session3
    '2019-01-14-15-07-21_1' : [4]         # ys_session1
}


# in the beginning experiments, not everyone started with 800 initial dwell time

dict_dwellTimeOrig_not800 = {
    "2019-01-16-15-51-13_2" : 600, # no_session1
    "2019-01-16-15-18-0_1" : 600, # no_session1
    "2019-01-16-15-43-8_1" : 100, # af_session1
    "2019-01-16-16-36-17_1stPart_2" : 100, # af_session1
    "2019-01-16-17-00-12_2ndPart_2" : 100, # af_session1
    "2019-01-17-15-03-40_1" : 100, # af_session2
    "2019-01-17-15-27-20_1stPart_2" : 0, # af_session2
    "2019-01-17-16-03-27_2ndPart_2" : 100, # af_session2
    "2019-01-14-15-07-21_1" : 500, # ys_session1
    "2019-01-14-15-25-55_2" : 300, # ys_session1
    "2019-01-16-15-18-50_1stpart_1" : 200, # ys_session2
    "2019-01-16-15-42-51_2ndPart_1" : 100, # ys_session2
    "2019-01-16-15-59-55_2" : 100, # ys_session2
    "2019-01-17-15-05-1_1" : 100, # ys_session3
    "2019-01-17-15-31-12_2" : 100 # ys_session3
}


# list of all things that should be present when computing effective time
list_keysToBeCounted = ['Comma', 'BackOne', 'BackMany', 'SpaceBar']

# some sessions do not have gaze data
dict_noGazeData = {
    '2019-01-16-17-00-12_2ndPart_2' : 'no gaze data', # af_session2
    '2019-01-17-15-31-12_2' : 'no gaze data', #ys_session2
    '2019-01-30-11-57-3_2' : 'no gaze data' # ys_session4
}


# In[ ]:


dict_keySelection_ReadingTrials = {
    "2019-01-16-15-42-51_2ndPart_1" : [0, 1, 3, 5], # ys_session2 
}

dict_keySelection_WritingTrials = {
    "2019-01-16-15-42-51_2ndPart_1" : [0, 1, 2, 5], # ys_session2   
}

# normally, reading part of trial ends when people look at the keyboardwithphrases. For some trials, this is not done,
# as the reading is done, and the trial is accidentally skipped, and written in the next trial. Here, the trial number 
# given will have the reading time ending as sleep, and not keyboard with phrases. 
dict_keyboardNotChange_ReadingTrials = {
    "2019-01-16-15-42-51_2ndPart_1" : 0, # ys_session2 
}

dict_keySelection_firstSleepNotCounted = {
    "2019-01-28-14-50-41_2" : (0, 2), # bh1_session1 -- 3rd sleep activation to be counted
    "2019-02-19-10-56-43_2ndPart_1" : 2  # mn_session3 -- 3rd sleep activation is to be counted
}


# In[ ]:


TimeDwellOrig = 800
TimeFixation = 300


# In[ ]:





# In[ ]:





# ## Clean input files

# In[ ]:


def FixUserKeys(UserKeys_Old):
    # Fix the situation where comma has divided decimals into separate columns
    
    Column_beforeDecimal = [item[2] for item in UserKeys_Old]
    Column_afterDecimal = [item[3] if len(item)>3 else '00' for item in UserKeys_Old]
    
    UserKeys_ProgressPercent = [float(Column_beforeDecimal[i]+'.'+ Column_afterDecimal[i]) for i in 
                                range(0, len(Column_beforeDecimal))]
    UserKeys_Times = [item[0] for item in UserKeys_Old]
    UserKeys_Keys = [item[1] for item in UserKeys_Old]
    
    UserKeys_New = [[UserKeys_Times[ind], UserKeys_Keys[ind], UserKeys_ProgressPercent[ind]] for ind in 
                    range(0, len(UserKeys_ProgressPercent))]
    
    #UserKeys_New = np.concatenate((UserKeys_Times, UserKeys_Keys, UserKeys_ProgressPercent), axis = 0)
    
    
    return UserKeys_New
        


# In[ ]:


def FixScratchPad(ScratchPad_Old):
    # Fix the situation where comma has divided decimals into separate columns
    
    ScratchPad_Times = [item[0] for item in ScratchPad_Old]
    
    ScratchPad_Phrases = list()
    
    # loop to combine phrases divided by commas
    ScratchPadInd = -1 
    while ScratchPadInd < len(ScratchPad_Old)-1:
        ScratchPadInd = ScratchPadInd + 1
        commasInPhrase = len(ScratchPad_Old[ScratchPadInd])-2
        if commasInPhrase < 1:
            #print(ScratchPad_Old[ScratchPadInd][1])
            ScratchPad_Phrases.append(ScratchPad_Old[ScratchPadInd][1])
            continue
        scratchPadPhrase = ScratchPad_Old[ScratchPadInd][1]
        for phraseJoinNr in range(1, commasInPhrase+1):
            scratchPadPhrase = scratchPadPhrase + ', ' + ScratchPad_Old[ScratchPadInd][1+phraseJoinNr]
        
        ScratchPad_Phrases.append(scratchPadPhrase)
            
        
    ScratchPad_New = [[ScratchPad_Times[ind], ScratchPad_Phrases[ind]] for ind in 
                    range(0, len(ScratchPad_Times))]
    
    #UserKeys_New = np.concatenate((UserKeys_Times, UserKeys_Keys, UserKeys_ProgressPercent), axis = 0)
    
    #print(ScratchPad_New)
    return ScratchPad_New


# In[ ]:


def FixKeysSelected(KeysSelected_Old):
    # Fix the situation where comma has divided decimals into separate columns
    
    KeysSelected_New = list()
    
    # loop to combine phrases divided by commas
    KeysSelectedInd = -1 
    while KeysSelectedInd < len(KeysSelected_Old)-1:
        KeysSelectedInd = KeysSelectedInd + 1
        
        if KeysSelected_Old[KeysSelectedInd][1].count(',') > 0:
            
            keys_split = KeysSelected_Old[KeysSelectedInd][1].split("\r\n")
            del keys_split[0]
            del keys_split[-1]
            
            keys_split = [key.split(',') for key in keys_split]
            
            KeysSelected_New.extend(keys_split)
        else:
            KeysSelected_New.append(KeysSelected_Old[KeysSelectedInd])
        
    
    return KeysSelected_New


# In[ ]:





# In[ ]:


def findAndRemoveTrials(session_name, dictionary_saved, trials, replacingList):
    # function to check the session_name in the dictionary_saved and remove those trials from the dictionary_trial
    
    if session_name in dictionary_saved:
        index_list = dictionary_saved[session_name]
    else:
        index_list = replacingList
    
    
    

    if index_list:
        if type(trials) == list:
            for index in sorted(index_list, reverse=True):
                del trials[index]
                
        else:
            for index in sorted(index_list, reverse=True):
                del trials['start'][index]
                del trials['end'][index]
            
    return trials    


# In[ ]:


def gazeConvert2ColumnsTo1(GazeLog, columnIndwValidity_list, indValidity):
    # function to convert pupilsizes from 2 columns for every pupil due to comma use instead of decimal, 
    # to proper pupil sizes
    
    #columnInd_list = [joinColumn1_1, joinColumn1_2, joinColumn2_1, joinColumn2_2]
    
    # number of columns in the final dictionary
    nColumns = int(len(columnIndwValidity_list)/2)
    
    # dictionary of columns that are to be joined later
    columns_beforeDecimal = dict()
    columns_afterDecimal = dict()
    
    # dictionary of joined columns
    columnsFinal = dict()
    
    # dictionary to find and equalize missing values in every column
    missingVal_column = dict()
    missingVal = list()
    
    # find correct index of validity column to be used, to find the actual columns relative to that
    columnsValidity_inUse = list()
    
    for ind, row in enumerate(GazeLog):
        #print(ind)
        #print(sorted(list(np.where(np.array(row) == 'Valid')[0])+list(np.where(np.array(row)=='Invalid')[0]))[indValidity])

        columnsValidity = (sorted(list(np.where(np.array(row) == 'Valid')[0])+list(np.where(np.array(row)=='Invalid')[0]))[indValidity])
        columnsValidity_inUse.append(int(columnsValidity))
    
    columnsValidity_inUse = np.array(columnsValidity_inUse)
    
    columnInd_list = [[columnsValidity_inUse+i] for i in columnIndwValidity_list]
    
    
    for ind in range(0, nColumns):
        
        dict_name = 'column' + str(ind+1)
        columnsFinal[dict_name] = list()
        columns_afterDecimal[dict_name] = list()
                
        #for indItem, item4 in enumerate(GazeLog):
        #    if 'Invalid' not in item4:
        #        if columnInd_list[2*ind+1][0][indItem] < len(item4):
        #            columns_afterDecimal[dict_name].append(item4[columnInd_list[2*ind+1][0][indItem]])
        #        else:
        #            columns_afterDecimal[dict_name].append('0')
        #    else:
        #        columns_afterDecimal[dict_name].append('nan')
        
                
        columns_beforeDecimal[dict_name] = [item4[columnInd_list[2*ind][0][indItem]] if 'Invalid' not in item4 else 'nan' for indItem, item4 in enumerate(GazeLog)]
        columns_afterDecimal[dict_name] = [item4[columnInd_list[2*ind+1][0][indItem]] if 'Invalid' not in item4 and columnInd_list[2*ind+1][0][indItem] < len(item4) else 'nan' for indItem, item4 in enumerate(GazeLog)]

        
        for i in range(0, len(columns_beforeDecimal[dict_name])):
            if 'Valid' not in columns_beforeDecimal[dict_name][i] and 'Valid' not in columns_afterDecimal[dict_name][i]:
                if 'nan' not in columns_beforeDecimal[dict_name][i] and 'nan' not in columns_afterDecimal[dict_name][i]:
                    if float(columns_afterDecimal[dict_name][i]) > 0: 
                        columnsFinal[dict_name].append(float(columns_beforeDecimal[dict_name][i]+'.'+columns_afterDecimal[dict_name][i]))
                    else:
                        columnsFinal[dict_name].append(np.nan)
                else:
                    columnsFinal[dict_name].append(np.nan)
            else:
                # Rarely, the pupil size is a whole number
                columnsFinal[dict_name].append(np.nan) # we will ignore the row, since there is no way of automatically knowing which - 
                # right or left eye has whole number pupil size
    
        missingVal_column[dict_name] = np.argwhere(np.isnan(columnsFinal[dict_name]))
        missingVal_column[dict_name] = list(itertools.chain.from_iterable(missingVal_column[dict_name])) # flatten the list
        
        missingVal.extend(missingVal_column[dict_name])
        
        
    
    missingVal = sorted(set(missingVal))
    
    # if one of the columns are nan, the other one is converted too
    for column in range(0, nColumns):
        dict_name = 'column' + str(column+1)
        for ind in missingVal:
            if ind < len(columnsFinal[dict_name]):
                columnsFinal[dict_name][ind] = np.nan
                
    
    
    
    #print(len(columnsFinal['column1']), len(columnsFinal['column2']))
    
    
    

    return columnsFinal


# ## Time list/data/formatting

# In[ ]:


def ComputeDwellTime(userKeys, full_path):
    # modify userKeys to include a column of time instead of progress pct, which is dependent on the then dwell time
    
    TimeDwellOrig = 800
    
    # session name
    session_folder_name = full_path.split('\\')[-1]
    
    if session_folder_name in dict_dwellTimeOrig_not800:
        TimeDwellOrig = dict_dwellTimeOrig_not800[session_folder_name]
    
    #print(TimeDwellOrig)
    
    timeDwell = TimeDwellOrig
    nKey = -1
    for key in userKeys:
        nKey = nKey + 1
        #print(key[1])
        if key[1] == 'IncreaseDwellTime':
            if float(key[2]) == 1:
                timeDwell = timeDwell + 100
        elif key[1] == 'DecreaseDwellTime':
            #print(key[2])
            if float(key[2]) == 1:
                timeDwell = timeDwell - 100
        else:
            userKeys[nKey].append(str(float(key[2])*timeDwell))
    
    return userKeys


# In[ ]:


# This function will return the datetime in items which is the closest to the date pivot
def nearestTimePoint(dates, date):
    
    for d in dates:
        if d <= date:
            nearestTP = d
        else:
            continue
    try: 
        nearestTP
        nearestTPind = dates.index(nearestTP)
    except:
        nearestTP = 0
        nearestTPind = -1
        
    return nearestTP, nearestTPind


# In[ ]:


# function to convert list of date and time into datetime format list
def timeConversion(timeStrList):
    timeList = list()
    for time in timeStrList:
        time1, t1, t2 = time.partition('+')
        timeList.append(datetime.datetime.strptime(re.sub('[:.T]','-',time1[:-1]), "%Y-%m-%d-%H-%M-%S-%f"))
    return timeList


# ## Trial timing from UserKeys

# In[ ]:




def OptiKeyTypingTime(UserKeys):
    import datetime
    import re
    re.compile
    TimeTyping = dict()
    
    time1, t1, t2 = UserKeys[0][0].partition('+')
    startTime = datetime.datetime.strptime(re.sub('[:.T]','-',time1[:-1]), "%Y-%m-%d-%H-%M-%S-%f")
    
    time2, t1, t2 = UserKeys[-1][0].partition('+')
    endTime = datetime.datetime.strptime(re.sub('[:.T]','-',time2[:-1]), "%Y-%m-%d-%H-%M-%S-%f")
    
    TimeTyping['startTime'] = startTime
    TimeTyping['endTime'] = endTime
    
    return TimeTyping


# In[ ]:


def timeTypingStart(userKeys):
    # From the user keys, find when the user actually starts typing, after having looked at the phrase and all the other 
    # function keys
    
    timeTypingStartInd = 0
    
    timeTypingStartIndList = list()
            
    timeUserTimeInd = 0
    
    ind = 0
    # Get start time of first trial
    
    while ind < len(userKeys):
        #print(len(userKeys[ind][1]))
        if len(userKeys[ind][1]) > 1:
            ind = ind + 1
        else:
            timeTypingStartInd = ind
            timeTypingStartIndList.append(ind)
            break
    
    #print(timeTypingStartInd)
    # Get every next phrase start timings
    while ind < len(userKeys):
        
        if userKeys[ind][1] == 'NextPhrase' and float(userKeys[ind][2]) == 1:
            
            #timeTypingStartIndList.append(ind+1)
            for ind2 in range(ind+1, len(userKeys)):
                if len(userKeys[ind2][1]) > 1:
                    ind = ind + 1
                    continue
                elif userKeys[ind2][1] == 'NextPhrase' and float(userKeys[ind][2]) == 1:
                    ind = ind + 1
                    continue
                else:
                    ind = ind2
                    timeTypingStartIndList.append(ind)
                    break
                    
        else:
            ind = ind + 1
            
    #print(timeTypingStartIndList)
    
    return timeTypingStartIndList


# ## Trial timing from Keys Selected and pupil data

# In[1]:


def FindExptStartEndTimes(KeysSelected, timeTyping, full_path):
    # function to find start and end of tasks in experiments
    
    
    
    # session name
    session_folder_name = full_path.split('\\')[-1]
    
    timeTrialDict = dict()
    timeTrialDict = {'start': [],
                    'end':[]}
    
    nTrial = -1
    
    
    
    for keys in KeysSelected:
        
            
        
        if keys[1] == 'NextPhrase':
            nTrial = nTrial + 1
            time1, t1, t2 = keys[0].partition('+')
            endTimeTrial = datetime.datetime.strptime(re.sub('[:.T]','-',time1[:-1]), "%Y-%m-%d-%H-%M-%S-%f")
            
            if nTrial != 0:
                # print('end: ', endTimeTrial)
                #print('')
                timeTrialDict['end'].append(endTimeTrial)
            
            
            # add 5s for the start time of the next phrase
            startTimeTrial = endTimeTrial + datetime.timedelta(seconds=5)
            
            #print('start: ', startTimeTrial)
            timeTrialDict['start'].append(startTimeTrial)
        
    del timeTrialDict['start'][-1]
    
    
    # remove the extra selections of NewPhrase at the end of some sessions
    replacingList = []
    timeTrialDict = findAndRemoveTrials(session_folder_name, dict_keySelectionOfNextPhrase, timeTrialDict, replacingList)
    
    timeTrialDict_copy = copy.deepcopy(timeTrialDict)
    
    # separate the reading and writing trials for some participants who read in the actual trial, but write in the next
    # trial
    if session_folder_name in dict_keySelection_ReadingTrials:
        # check the reading and writing separate dictionaries
        print('reading and writing sessions are separate')
        
        #print(len(timeTrialDict['start']))
        # writing trials - 
        timeTrialDict_writing = findAndRemoveTrials(session_folder_name, dict_keySelection_WritingTrials, timeTrialDict, replacingList)
        #print(len(timeTrialDict_copy['start']))
        
        # reading trials
        timeTrialDict_reading = findAndRemoveTrials(session_folder_name, dict_keySelection_ReadingTrials, timeTrialDict_copy, replacingList)
    else:
        # some participants skip some sentences, and then it affects the scoreQuestions too. Remove the skipped sentences or 
        # remove the score questions 
        # for these participants, reading and writing trials are the same
        
        #print('same reading and writing trials')
        scoreQuestions = [0, 1, 3, 5, 7, 9, 11]
        timeTrialDict = findAndRemoveTrials(session_folder_name, dict_keySelectionNotCompleted, timeTrialDict, scoreQuestions)
        
        # most of the skipped sentences are removed, but for those that are not removed
        timeTrialDict_writing = findAndRemoveTrials(session_folder_name, dict_keySelection_phraseStim, timeTrialDict, replacingList)
        
        timeTrialDict_reading = timeTrialDict_writing 
        
    
          
    return timeTrialDict_reading, timeTrialDict_writing


# In[ ]:


def FindReadingPartsOfTrial_inKeysSelected(EventTrials_reading, KeysSelected_new, full_path, userKeys):
    # find the reading end in the pupil size data
    
    # session name
    session_folder_name = full_path.split('\\')[-1]
    
    KeysSelected_timeStr = [key[0] for key in KeysSelected_new]
    KeysSelected_time = timeConversion(KeysSelected_timeStr)
    
    KeysSelected_keys = [key[1] for key in KeysSelected_new]
    
    userKeysTimeStr = [key[0] for key in userKeys]
    userKeysTime = timeConversion(userKeysTimeStr)
    
    EventReading = dict()    
    EventReading['start'] = list()
    EventReading['end'] = list()
    
    EventReading_index = dict()    
    EventReading_index['start'] = list()
    EventReading_index['end'] = list()
    
    for ind, startTrialTime_afterCoolDown in enumerate(EventTrials_reading['start']):
        
        startTrialTime, startTrialInd = nearestTimePoint(KeysSelected_time, startTrialTime_afterCoolDown)
        
        
        EventReading['start'].append(KeysSelected_time[startTrialInd]+datetime.timedelta(seconds=5)) # start time needs 
        #to start 5s later, which is when the phrase is visible
        EventReading_index['start'].append(startTrialInd) 
        
        
        #print(ind, EventReading['start'][-1], EventReading_index['start'][-1])
        
        endTrialTime = EventTrials_reading['end'][ind]
        endTrialInd = KeysSelected_time.index(endTrialTime)
        
        keysSelected_trial = KeysSelected_keys[startTrialInd:endTrialInd]
        
        
        for i, key in enumerate(keysSelected_trial):
            if len(key) == 1:
                endReading_keyInd = startTrialInd + keysSelected_trial.index(key) - 1
                break
                    
        endReading_keyTime = KeysSelected_time[endReading_keyInd]
        
        
        userKeyTime, userKeyInd = nearestTimePoint(userKeysTime, KeysSelected_time[endReading_keyInd+1])
        
        # remove the dwell time from end of selecting the first key (letter/number)
        EventReading['end'].append(KeysSelected_time[endReading_keyInd+1]-datetime.timedelta(milliseconds=                                    float(userKeys[userKeyInd][-1])))
        
        EventReading_index['end'].append(KeysSelected_time[endReading_keyInd+1])
        
        
    return EventReading, EventReading_index


# In[ ]:


def FindWritingPartsOfTrial_inKeysSelected(EventTrials_writing, KeysSelected_new, EventReading):
    
    KeysSelected_timeStr = [key[0] for key in KeysSelected_new]
    KeysSelected_time = timeConversion(KeysSelected_timeStr)
    
    KeysSelected_keys = [key[1] for key in KeysSelected_new]
    
    EventWriting = dict()    
    EventWriting['start'] = list()
    EventWriting['end'] = list()
    
    EventWriting_index = dict()    
    EventWriting_index['start'] = list()
    EventWriting_index['end'] = list()
    
    for ind, startTrialTime_afterCoolDown in enumerate(EventTrials_writing['start']):
        
        startTrialTime, startTrialInd = nearestTimePoint(KeysSelected_time, startTrialTime_afterCoolDown)
        
        endTrialTime = EventTrials_writing['end'][ind]
        endTimeReading = EventReading['end'][ind]
        
        
        # for some participants, reading and writing trials are different. So their writing times will not be the end of 
        # the reading time.
        # Regardless, the writing time should start later than when the reading time ends.
        # We choose the starting time for writing as the one that is later than the start time from writing trials
        # and end time from reading trials
        
        if startTrialTime > endTimeReading:
            EventWriting['start'].append(startTrialTime)
            EventWriting_index['start'].append(startTrialInd)
        else:
            EventWriting['start'].append(endTimeReading)
            endTimeReading_keys, endTimeReading_ind = nearestTimePoint(KeysSelected_time, endTimeReading)
            EventWriting_index['start'].append(KeysSelected_time.index(endTimeReading_keys))
        
        EventWriting['end'].append(endTrialTime)
        EventWriting_index['end'].append(KeysSelected_time.index(endTrialTime))
        
        #print(ind)
        #print('writing: ', EventWriting['start'][ind], EventWriting['end'][ind])
        
    
    
    return EventWriting, EventWriting_index     


# In[ ]:


def EventPartsFromPupilData(EventTimeInKeys, PupilSize_df, full_path):
    
    # session name
    session_folder_name = full_path.split('\\')[-1]
    
    EventTime = dict()    
    EventTime['start'] = list()
    EventTime['end'] = list()
    
    EventIndex = dict()    
    EventIndex['start'] = list()
    EventIndex['end'] = list()
    
    EventBaseline_startKeyTime = list()
    
    for ind, eventStartInKeys in enumerate(EventTimeInKeys['start']):
        
        eventStartTime, eventStartInd = nearestTimePoint(PupilSize_df['timeStamp'].tolist(), eventStartInKeys)
        eventEndTime, eventEndInd = nearestTimePoint(PupilSize_df['timeStamp'].tolist(), EventTimeInKeys['end'][ind])
        
        # reading start is the same as trial start
        EventTime['start'].append(eventStartTime)
        EventIndex['start'].append(eventStartInd)
        
        EventTime['end'].append(eventEndTime)
        EventIndex['end'].append(eventEndInd)
        
        
    return EventTime, EventIndex    


# # Filter Pupil size only

# In[ ]:


def FilterPupilSize(GazeLog, TimeTyping, subjectAndSessionName):
    # function that uses the list of start and end trial times to find the pupil sizes for those trials and plots them
    
    # first create a list of times in gaze log
    timeStrGazeLog = [item3[0] for item3 in GazeLog]
    # convert the list of strings to datetime formats
    timeGazeLog = timeConversion(timeStrGazeLog)
    
    # internal time, to depict seconds
    timeInternalGazeLog = [float(item3[1]) for item3 in GazeLog]
    
    # extract pupil sizes in decimals from the strange 2 columns for every pupil
    pupil_indWrtValidityL = [1, 2]
    pupil_validityL = 4
    pupilLogL_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityL, pupil_validityL)
    
    pupil_indWrtValidityR = [1, 2]
    pupil_validityR = 5
    pupilLogR_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityR, pupil_validityR)
    
    
    # reduce the data to start and end of typing time
    timeTyping_start, timeTyping_startInd = nearestTimePoint(timeGazeLog, TimeTyping['startTime'])
    timeTyping_end, timeTyping_endInd = nearestTimePoint(timeGazeLog, TimeTyping['endTime'])
    
    pupilLogL_wDefinedTime = pupilLogL_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    pupilLogR_wDefinedTime = pupilLogR_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    
    timeGazeLog_wDefinedTime = timeGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    
    timeInS_GazeLog_wDefinedTime = timeInternalGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    timeInS_Difference = [(t - s)/1000000 for s, t in zip(timeInS_GazeLog_wDefinedTime, timeInS_GazeLog_wDefinedTime[1:])]
    timeInS_Difference.insert(0, 0)
    
    
    #timeInS = [sum(timeInS_Difference[:i]) for i, v in enumerate(timeInS_Difference)]
    
    pupilData_df, interpolated_items = filterBlinks(pupilLogL_wDefinedTime, pupilLogR_wDefinedTime, timeGazeLog_wDefinedTime)
    
    
    
    
    #timeGazeLog_plot = np.arange(0, timeInS[-1], 1/90)
    
    #plotPupilSize_checkFilter(pupilData_df, pupilLogL_wDefinedTime, blinkStartAndEnd, 'blink removal', subjectAndSessionName)
    
    
    
    
    pupilData_df_hampel = dict()
    
    pupilData_df_hampel = pupilData_df.copy()
    pupilData_df_hampel['pupilLeft'] = hampel(pupilData_df['pupilLeft'], 25, 3, False)
    pupilData_df_hampel['pupilRight'] = hampel(pupilData_df['pupilRight'], 25, 3, False)
        
    
        
    return pupilData_df_hampel, interpolated_items


# ## Filter pupil size with weighted pupil mean

# In[ ]:


def filterBlinks(pupilDataL, pupilDataR, timeInDatetime):
    # filter any blinks and nan values lasting around 250ms (on average)
    # first the single nan occurances are replaced with mean of the values on either sides, 
    # as they are assumed to be from hardware problems
    # for the rest of the blinks, 250ms before and after the nan values are interpolated with a linear function
    # returns a dataframe with pupil size, and timestamp
    # http://faculty.washington.edu/chudler/facts.html
   
    
    # create a dataframe from the pupilsize and time
    pupilData_df = pd.DataFrame(list(zip(timeInDatetime, pupilDataL, pupilDataR)), columns=['timeStamp', 'pupilLeft', 'pupilRight'])
    
    # blink is every nan value in the range of 100-400ms 
    # 250 ms (22 samples) before and after the blink will also be removed
    extraBlinkSamples = 18
    
    
    #pupilData_woSingleMissingData = pupilData.copy()
    #timeList_woSingleMissingData = timeInDatetime.copy()
    #timeInS_woSingleMissingData = timeInS_Trial[-1]
    
    # in case of single missing data, that are due to hardware error, replace with the mean of the pupil size before and
    # after the nan value
    # missing values will be the same for left and right pupil
    missingVal_Single = np.argwhere(np.isnan(pupilDataL))
    missingVal_Single = list(itertools.chain.from_iterable(missingVal_Single)) # flatten the list 
    
    
    
    # if no blinks present, return the data
    if len(missingVal_Single) == 0:
        interpolatedNan_bool = np.array([False]*len(pupilData_df['pupilLeft']))
        return pupilData_df, interpolatedNan_bool
    
    # find the index and values to replace for single nan values
    pupilData_tuples_replaceSingleNan_left = [(val, np.mean([pupilDataL[val-1], pupilDataL[val+1]])) for i, val in enumerate(missingVal_Single) if (val != 0 and val != (len(pupilDataL)-1)) if not np.isnan(pupilDataL[val-1]) and not np.isnan(pupilDataL[val+1])]
    pupilData_tuples_replaceSingleNan_right = [(val, np.mean([pupilDataR[val-1], pupilDataR[val+1]])) for i, val in enumerate(missingVal_Single) if (val != 0 and val != (len(pupilDataR)-1)) if not np.isnan(pupilDataR[val-1]) and not np.isnan(pupilDataR[val+1])]
    
    
    interpolatedNan_bool = np.array([True if ind in dict(pupilData_tuples_replaceSingleNan_left) else False for ind, val in enumerate(pupilDataL)])
    missingData_single = interpolatedNan_bool
    
    # replace the single nan values with the mean of the pupil size on either sides
    indList = -1
    for ind, val in pupilData_tuples_replaceSingleNan_left:
        indList = indList + 1
        pupilData_df.iloc[ind, pupilData_df.columns.get_loc('pupilLeft')] = val
        pupilData_df.iloc[ind, pupilData_df.columns.get_loc('pupilRight')] = pupilData_tuples_replaceSingleNan_right[indList][1]
        
    
    # again, find the nan values in the pupil size
    # the list missingVal_SingleDifference contains the index of the first blink, followed by the difference in the index
    # to the next nan value
    
    
    
    # find the nan values again from pupilData['pupilLeft']
    missingVal_Rest_trueFalse = pupilData_df['pupilLeft'].isnull()
    missingVal_Rest = [i for i, x in enumerate(missingVal_Rest_trueFalse) if x]
    
    # if no blinks left, return the current pupilData
    if len(missingVal_Rest) == 0:
        return pupilData_df, interpolatedNan_bool

    
    # in the blinks left, find when the blinks start by finding a difference in the consecutive values of the indices
    missingVal_RestDifference = [t - s for s, t in zip(missingVal_Rest, missingVal_Rest[1:])]
    missingVal_RestDifference.insert(0, missingVal_Rest[0])
    
    blinkStart_tupleList = [(ind, sum(missingVal_RestDifference[0:ind+1])) for ind, val in enumerate(missingVal_RestDifference) if val != 1]
    
    blinkStart_tupleList_wLength = list()
    
    # create a list of tuples of blink start index and the length of the blink
    ind = -1
    blinkLengthSum = 0
    for blink_ind, blinkStartInd in blinkStart_tupleList:
        ind = ind + 1
        if ind != len(blinkStart_tupleList) - 1:
            
            blinkLength = blinkStart_tupleList[ind+1][0]-blink_ind
            blinkLengthSum = blinkLengthSum + blinkLength
            
            blinkStart_tupleList_wLength.append(tuple((blinkStartInd, blinkLength)))
        else:
            # for the last blink -- all blink lengths summed and subtracted from the length of the list
            # missingVal_RestDifference 
            blinkLength = len(missingVal_RestDifference)-blinkLengthSum
            blinkStart_tupleList_wLength.append(tuple((blinkStartInd, blinkLength)))
     
    
    
    
    # create a vector with True if a blink was at the position
    samplingFrequency = 90
    blinkLengthMax = np.ceil(0.5*samplingFrequency)
    blinkLengthMin = np.ceil(0.075*samplingFrequency)
    blinkStart_tupleList_wLength_allMissingData = [(blinkStart, blinkLength) for blinkStart, blinkLength in blinkStart_tupleList_wLength]    
    blinkStart_tupleList_wLength_maxBlinkLength = [(blinkStart, blinkLength) for blinkStart, blinkLength in blinkStart_tupleList_wLength if blinkLength<blinkLengthMax and blinkLength>blinkLengthMin]
    
    missingData_blinks_bool = np.array([False]*len(pupilData_df['pupilLeft']))
    for blinkStart, blinkLength in blinkStart_tupleList_wLength_maxBlinkLength:
        blinkIndices = np.arange(blinkStart, blinkStart + blinkLength).astype(int)
        missingData_blinks_bool[blinkIndices] = True
        
    
    
    missingData_blinks_boolAll = np.array([False]*len(pupilData_df['pupilLeft']))
    for blinkStart, blinkLength in blinkStart_tupleList_wLength:
        blinkIndices = np.arange(blinkStart, blinkStart + blinkLength).astype(int)
        missingData_blinks_boolAll[blinkIndices] = True
        
    missingDataOverall_bool = missingData_single + missingData_blinks_boolAll
    
    # add to vector with start and end of tuple
    #beforeAfterNan = [False]*len(pupilData_df['pupilLeft'])
    #for blinkStart, blinkLength in blinkStart_tupleList_wLength:
    #    beforeAfterNan[blinkStart] = True
    #    beforeAfterNan[blinkStart+blinkLength] = True
    #    #print('start and end points: ', pupilData_df['timeStamp'][blinkStart], pupilData_df['timeStamp'][blinkStart + blinkLength])
    
    
    # create lists with start and end values for the blinks, based on blinkStart_tupleList_wLength, regardless of the blink length
    blink_missingData_startList = [blinkStartInd - extraBlinkSamples if (blinkStartInd - extraBlinkSamples) > 0 else 0 for blinkStartInd, blinkLength in blinkStart_tupleList_wLength]
    blink_missingData_endList = [blinkStartInd + blinkLength + extraBlinkSamples if (blinkStartInd + blinkLength + extraBlinkSamples) < (len(pupilData_df['pupilLeft'])-1) else (len(pupilData_df['pupilLeft'])-1) for blinkStartInd, blinkLength in blinkStart_tupleList_wLength]
    # create a list of tuples from the start and end points of the blink
    blink_missingData_startEndTuple = [(blinkStart, blink_missingData_endList[ind]) for ind, blinkStart in enumerate(blink_missingData_startList)] 
    
    
    # check if blinks need to be combined - blinksCombine is a list of list of 2 elements, the index of the blinks that should be combined
    blinksCombine = [[ind, ind+1] for ind, blink in enumerate(blink_missingData_startEndTuple[0:-1]) if blink[1] > blink_missingData_startEndTuple[ind+1][0]]
        
    if blinksCombine:
        # combine blinks that need to be combined - if multiple consecutive blinks need to be removed: eg - [1, 2], [2, 3] 
        # are included in the blinksCombine, the combined version should be [1, 3] 
        blinksCombineFinal = list()
        ind = -1
        while ind < len(blinksCombine)-2:
            
            ind = ind + 1
            blinkCombining = blinksCombine[ind]
            blinksCombineFinal.append(blinkCombining)
            while ind < len(blinksCombine)-2 and blinkCombining[1] == blinksCombine[ind+1][0]:
                # change the ending of the last added blink of blinksCombineFinal
                blinksCombineFinal[-1][1] = blinksCombine[ind+1][1]
                ind = ind + 1
            
            
        if len(blinksCombine) == 1:
            blinksCombineFinal = blinksCombine.copy()
            
        
        if blinksCombine[-1][1] != blinksCombineFinal[-1][1]:
            if blinksCombine[-1][0] == blinksCombineFinal[-1][1]:
                blinksCombineFinal[-1][1] = blinksCombine[-1][1]
            else:
                blinksCombineFinal.append(blinksCombine[-1])
        
        
        
        
    #    for w, z in groupby(sorted(list(blinksCombine)), lambda x, y=itertools.count(): next(y)-x):
    #        group = list(z)
    #        blinksCombineFinal.append(tuple((group[0], group[-1])))
        
        for x in sorted(blinksCombineFinal, reverse=True):
            new_start = blink_missingData_startEndTuple[x[0]][0] 
            new_end = blink_missingData_startEndTuple[x[1]][1] 
            
            x_start = x[0]
            x_end = x[1]
            
            # delete also the blinkStart_tupleList_wLength, since it is going to be used to compute other metrics
            for blinkRemove in range(x[1], x[0]-1, -1):
                del blink_missingData_startEndTuple[blinkRemove]
            
            blink_missingData_startEndTuple.insert(x[0], tuple((new_start, new_end)))
    
    
    #blinkAndNonBlinkDurationList = [length/90 for start, length in blinkStart_tupleList_wLength]
    #timeInS_Trial_filter = timeInS_Trial[-1] - sum(blinkAndNonBlinkDurationList) 
    
    
    # remove blinks from data
    for blinkStart, blinkEnd in blink_missingData_startEndTuple:
        pupilData_df.loc[blinkStart:blinkEnd,'pupilLeft'] = np.nan
        pupilData_df.loc[blinkStart:blinkEnd,'pupilRight'] = np.nan
        replaceTrueList = range(blinkStart, blinkEnd+1, 1)
        interpolatedNan_bool[replaceTrueList] = True
    
    
    
    pupilData_df['pupilLeft'] = pupilData_df['pupilLeft'].astype(float).interpolate('linear', limit_direction = 'both')
    pupilData_df['pupilRight'] = pupilData_df['pupilRight'].astype(float).interpolate('linear', limit_direction = 'both')
    
    if pupilData_df.isnull().any().any():
        print('nan values in filtered data')
        #for i,val in enumerate(pupilData_filter[0:5000]):
        #    print(i, val, pupilData_woSingleMissingData[i])
        
    
    return pupilData_df, interpolatedNan_bool


# In[ ]:


def hampel(dvec, radius=5, nsigma=3, rem_nomed=False):

    # replace outliers with median values (hampel filter)
    
    mvec = pd.Series(dvec).rolling(radius*2+1, center=True, min_periods=radius).median()
    svec = 1.4862 * np.abs(dvec-mvec).rolling(radius*2+1, center=True, min_periods=radius).median()
    plonk = np.abs(dvec-mvec) > nsigma*svec
    dvec = np.array(dvec)
    dvec[plonk.tolist()] = mvec[plonk.tolist()]

    # remove "bad data" where we cannot calculate a median value due to already missing values
    if (rem_nomed):
        dvec[np.isnan(mvec)] = np.nan
    return dvec


# In[ ]:


def weightedMean(pupilData_df, window):
    
    # find the rolling standard deviations
    std_left_end = pupilData_df['pupilLeft'].rolling(window).std()[window-1:]
    std_right_end = pupilData_df['pupilRight'].rolling(window).std()[window-1:]
    
    std_left = [np.std(pupilData_df['pupilLeft'][0:i]) for i in range(1,window-1)]
    std_left.insert(0,0)
    
    
    std_right = [np.std(pupilData_df['pupilRight'][0:i]) for i in range(1,window-1)]
    std_right.insert(0,0)
    
    
        
    std_left.extend(std_left_end)
    std_right.extend(std_right_end)
    
    
    # if std_left is 0, the succeeding value is taken
    i = len(std_left)-1
    while i > -1:
        left = std_left[i]
        if left == 0:
            if i < len(std_left)-1:
                std_left[i] = std_left[i+1]
            else:
                std_left[i] = 2 # to give 0.5 for 1/std
        i = i - 1
    
    # if std_left is 0, the succeeding value is taken
    i = len(std_right) - 1
    while i > -1:
        right = std_right[i]
        if right == 0:
            if i < len(std_right)-1:
                std_right[i] = std_right[i+1]
                
            else:
                std_right[i] = 2 # to give 0.5 for 1/std
        i = i - 1
    
    # coefficients are 1/std
    coeff_left_notNormalized = 1/np.array(std_left)
    coeff_right_notNormalized = 1/np.array(std_right)
    
    coeff_left = coeff_left_notNormalized/(coeff_left_notNormalized+coeff_right_notNormalized)
    coeff_right_a = coeff_right_notNormalized/(coeff_left_notNormalized+coeff_right_notNormalized)
    coeff_right = 1-coeff_left
    
    #for i, x in enumerate(coeff_left):
    #    print(coeff_left_notNormalized[i], x, coeff_right_notNormalized[i], coeff_right[i])
    
    
    #coeff_left_notNormalized = [1/std if not np.isnan(std) and std!=0 else 0.5 for std in std_left]
    #coeff_right_notNormalized = [1/std if not np.isnan(std) and std!=0 else 0.5 for std in std_right]
    
    #coeff_left = np.array([left/(left+coeff_right_notNormalized[ind]) for ind, left in enumerate(coeff_left_notNormalized)])
    #coeff_right = 1-coeff_left
    
    pupilData_df['pupilMean_weighted'] = coeff_left*pupilData_df['pupilLeft'] + coeff_right*pupilData_df['pupilRight']
    pupilData_df['pupilMean'] = 0.5*pupilData_df['pupilLeft'] + 0.5*pupilData_df['pupilRight']
    
    #error = [np.abs(np.sum(pupilData_df['pupilLeft'][ind:ind+window-1])-np.sum(pupilData_df['pupilRight'][ind:ind+window-1]))/window for ind, left in enumerate(pupilData_df['pupilLeft'])]
    
    
    #print(len(pupilData_df['pupilLeft']), len(error))
    
    #print(error[0:20])
    #print(error[-20:])
    
    print(np.corrcoef(pupilData_df['pupilLeft'], pupilData_df['pupilRight']))
    
    corr = np.corrcoef(pupilData_df['pupilLeft'], pupilData_df['pupilRight'])[0][1]
    
    
    #error = np.abs(pupilData_df['pupilLeft'] - pupilData_df['pupilRight'])
    
    #plt.figure()
    #plt.plot(error)
    #plt.plot(pupilData_df['pupilLeft'])
    #plt.plot(pupilData_df['pupilRight'])
    
    
    
    
    return pupilData_df, corr


# In[ ]:


def FilterPupilSize_wWeightedMean(GazeLog, TimeTyping, subjectAndSessionName):
    # function that uses the list of start and end trial times to find the pupil sizes for those trials and plots them
    
    # first create a list of times in gaze log
    timeStrGazeLog = [item3[0] for item3 in GazeLog]
    # convert the list of strings to datetime formats
    timeGazeLog = timeConversion(timeStrGazeLog)
    
    # internal time, to depict seconds
    timeInternalGazeLog = [float(item3[1]) for item3 in GazeLog]
    
    # extract pupil sizes in decimals from the strange 2 columns for every pupil
    pupil_indWrtValidityL = [1, 2]
    pupil_validityL = 4
    pupilLogL_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityL, pupil_validityL)
    
    pupil_indWrtValidityR = [1, 2]
    pupil_validityR = 5
    pupilLogR_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityR, pupil_validityR)
    
    
    # reduce the data to start and end of typing time
    timeTyping_start, timeTyping_startInd = nearestTimePoint(timeGazeLog, TimeTyping['startTime'])
    timeTyping_end, timeTyping_endInd = nearestTimePoint(timeGazeLog, TimeTyping['endTime'])
    
    pupilLogL_wDefinedTime = pupilLogL_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    pupilLogR_wDefinedTime = pupilLogR_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    
    timeGazeLog_wDefinedTime = timeGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    
    timeInS_GazeLog_wDefinedTime = timeInternalGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    timeInS_Difference = [(t - s)/1000000 for s, t in zip(timeInS_GazeLog_wDefinedTime, timeInS_GazeLog_wDefinedTime[1:])]
    timeInS_Difference.insert(0, 0)
    
    
    #timeInS = [sum(timeInS_Difference[:i]) for i, v in enumerate(timeInS_Difference)]
    
    pupilData_df, interpolated_items = filterBlinks(pupilLogL_wDefinedTime, pupilLogR_wDefinedTime, timeGazeLog_wDefinedTime)
    
    
    
    
    #timeGazeLog_plot = np.arange(0, timeInS[-1], 1/90)
    
    #plotPupilSize_checkFilter(pupilData_df, pupilLogL_wDefinedTime, blinkStartAndEnd, 'blink removal', subjectAndSessionName)
    
    
    
    
    pupilData_df_hampel = dict()
    
    pupilData_df_hampel = pupilData_df.copy()
    pupilData_df_hampel['pupilLeft'] = hampel(pupilData_df['pupilLeft'], 25, 3, False)
    pupilData_df_hampel['pupilRight'] = hampel(pupilData_df['pupilRight'], 25, 3, False)
        
    
    # weighted mean of right and left pupils, based on their standard deviation 
    std_window = 25
    pupilDataMean_df, correlationRL = weightedMean(pupilData_df_hampel, std_window)
    
    
        
    return pupilDataMean_df, interpolated_items, correlationRL


# ## Filter pupil size with blink data

# In[ ]:


def filterBlinks_wBlinkData(pupilDataL, pupilDataR, timeInDatetime):
    # filter any blinks and nan values lasting around 250ms (on average)
    # first the single nan occurances are replaced with mean of the values on either sides, 
    # as they are assumed to be from hardware problems
    # for the rest of the blinks, 250ms before and after the nan values are interpolated with a linear function
    # returns a dataframe with pupil size, and timestamp
    # http://faculty.washington.edu/chudler/facts.html
   
    
    # create a dataframe from the pupilsize and time
    pupilData_df = pd.DataFrame(list(zip(timeInDatetime, pupilDataL, pupilDataR)), columns=['timeStamp', 'pupilLeft', 'pupilRight'])
    
    # blink is every nan value in the range of 100-400ms 
    # 250 ms (22 samples) before and after the blink will also be removed
    extraBlinkSamples = 18
    
    
    #pupilData_woSingleMissingData = pupilData.copy()
    #timeList_woSingleMissingData = timeInDatetime.copy()
    #timeInS_woSingleMissingData = timeInS_Trial[-1]
    
    # in case of single missing data, that are due to hardware error, replace with the mean of the pupil size before and
    # after the nan value
    # missing values will be the same for left and right pupil
    missingVal_Single = np.argwhere(np.isnan(pupilDataL))
    missingVal_Single = list(itertools.chain.from_iterable(missingVal_Single)) # flatten the list 
    
    
    
    # if no blinks present, return the data
    if len(missingVal_Single) == 0:
        interpolatedNan_bool = np.array([False]*len(pupilData_df['pupilLeft']))
        return pupilData_df, interpolatedNan_bool
    
    # find the index and values to replace for single nan values
    pupilData_tuples_replaceSingleNan_left = [(val, np.mean([pupilDataL[val-1], pupilDataL[val+1]])) for i, val in enumerate(missingVal_Single) if (val != 0 and val != (len(pupilDataL)-1)) if not np.isnan(pupilDataL[val-1]) and not np.isnan(pupilDataL[val+1])]
    pupilData_tuples_replaceSingleNan_right = [(val, np.mean([pupilDataR[val-1], pupilDataR[val+1]])) for i, val in enumerate(missingVal_Single) if (val != 0 and val != (len(pupilDataR)-1)) if not np.isnan(pupilDataR[val-1]) and not np.isnan(pupilDataR[val+1])]
    
    
    interpolatedNan_bool = np.array([True if ind in dict(pupilData_tuples_replaceSingleNan_left) else False for ind, val in enumerate(pupilDataL)])
    missingData_single = interpolatedNan_bool
    
    # replace the single nan values with the mean of the pupil size on either sides
    indList = -1
    for ind, val in pupilData_tuples_replaceSingleNan_left:
        indList = indList + 1
        pupilData_df.iloc[ind, pupilData_df.columns.get_loc('pupilLeft')] = val
        pupilData_df.iloc[ind, pupilData_df.columns.get_loc('pupilRight')] = pupilData_tuples_replaceSingleNan_right[indList][1]
        
    
    # again, find the nan values in the pupil size
    # the list missingVal_SingleDifference contains the index of the first blink, followed by the difference in the index
    # to the next nan value
    
    
    
    # find the nan values again from pupilData['pupilLeft']
    missingVal_Rest_trueFalse = pupilData_df['pupilLeft'].isnull()
    missingVal_Rest = [i for i, x in enumerate(missingVal_Rest_trueFalse) if x]
    
    # if no blinks left, return the current pupilData
    if len(missingVal_Rest) == 0:
        return pupilData_df, interpolatedNan_bool

    
    # in the blinks left, find when the blinks start by finding a difference in the consecutive values of the indices
    missingVal_RestDifference = [t - s for s, t in zip(missingVal_Rest, missingVal_Rest[1:])]
    missingVal_RestDifference.insert(0, missingVal_Rest[0])
    
    blinkStart_tupleList = [(ind, sum(missingVal_RestDifference[0:ind+1])) for ind, val in enumerate(missingVal_RestDifference) if val != 1]
    
    blinkStart_tupleList_wLength = list()
    
    # create a list of tuples of blink start index and the length of the blink
    ind = -1
    blinkLengthSum = 0
    for blink_ind, blinkStartInd in blinkStart_tupleList:
        ind = ind + 1
        if ind != len(blinkStart_tupleList) - 1:
            
            blinkLength = blinkStart_tupleList[ind+1][0]-blink_ind
            blinkLengthSum = blinkLengthSum + blinkLength
            
            blinkStart_tupleList_wLength.append(tuple((blinkStartInd, blinkLength)))
        else:
            # for the last blink -- all blink lengths summed and subtracted from the length of the list
            # missingVal_RestDifference 
            blinkLength = len(missingVal_RestDifference)-blinkLengthSum
            blinkStart_tupleList_wLength.append(tuple((blinkStartInd, blinkLength)))
     
    
    
    
    # create a vector with True if a blink was at the position
    samplingFrequency = 90
    blinkLengthMax = np.ceil(0.5*samplingFrequency)
    blinkLengthMin = np.ceil(0.075*samplingFrequency)
    blinkStart_tupleList_wLength_allMissingData = [(blinkStart, blinkLength) for blinkStart, blinkLength in blinkStart_tupleList_wLength]    
    blinkStart_tupleList_wLength_maxBlinkLength = [(blinkStart, blinkLength) for blinkStart, blinkLength in blinkStart_tupleList_wLength if blinkLength<blinkLengthMax and blinkLength>blinkLengthMin]
    
    missingData_blinks_bool = np.array([False]*len(pupilData_df['pupilLeft']))
    for blinkStart, blinkLength in blinkStart_tupleList_wLength_maxBlinkLength:
        blinkIndices = np.arange(blinkStart, blinkStart + blinkLength).astype(int)
        if blinkIndices[-1] > len(pupilDataL):
            blinkIndices = np.arange(blinkStart, len(pupilDataL)-1)
        missingData_blinks_bool[blinkIndices] = True
        
    
    
    missingData_blinks_boolAll = np.array([False]*len(pupilData_df['pupilLeft']))
    for blinkStart, blinkLength in blinkStart_tupleList_wLength:
        blinkIndices = np.arange(blinkStart, blinkStart + blinkLength).astype(int)
        if blinkIndices[-1] > len(pupilDataL):
            blinkIndices = np.arange(blinkStart, len(pupilDataL)-1)
        missingData_blinks_boolAll[blinkIndices] = True
        
    missingDataOverall_bool = missingData_single + missingData_blinks_boolAll
    
    # add to vector with start and end of tuple
    #beforeAfterNan = [False]*len(pupilData_df['pupilLeft'])
    #for blinkStart, blinkLength in blinkStart_tupleList_wLength:
    #    beforeAfterNan[blinkStart] = True
    #    beforeAfterNan[blinkStart+blinkLength] = True
    #    #print('start and end points: ', pupilData_df['timeStamp'][blinkStart], pupilData_df['timeStamp'][blinkStart + blinkLength])
    
    
    # create lists with start and end values for the blinks, based on blinkStart_tupleList_wLength, regardless of the blink length
    blink_missingData_startList = [blinkStartInd - extraBlinkSamples if (blinkStartInd - extraBlinkSamples) > 0 else 0 for blinkStartInd, blinkLength in blinkStart_tupleList_wLength]
    blink_missingData_endList = [blinkStartInd + blinkLength + extraBlinkSamples if (blinkStartInd + blinkLength + extraBlinkSamples) < (len(pupilData_df['pupilLeft'])-1) else (len(pupilData_df['pupilLeft'])-1) for blinkStartInd, blinkLength in blinkStart_tupleList_wLength]
    # create a list of tuples from the start and end points of the blink
    blink_missingData_startEndTuple = [(blinkStart, blink_missingData_endList[ind]) for ind, blinkStart in enumerate(blink_missingData_startList)] 
    
    
    # check if blinks need to be combined - blinksCombine is a list of list of 2 elements, the index of the blinks that should be combined
    blinksCombine = [[ind, ind+1] for ind, blink in enumerate(blink_missingData_startEndTuple[0:-1]) if blink[1] > blink_missingData_startEndTuple[ind+1][0]]
        
    if blinksCombine:
        # combine blinks that need to be combined - if multiple consecutive blinks need to be removed: eg - [1, 2], [2, 3] 
        # are included in the blinksCombine, the combined version should be [1, 3] 
        blinksCombineFinal = list()
        ind = -1
        while ind < len(blinksCombine)-2:
            
            ind = ind + 1
            blinkCombining = blinksCombine[ind]
            blinksCombineFinal.append(blinkCombining)
            while ind < len(blinksCombine)-2 and blinkCombining[1] == blinksCombine[ind+1][0]:
                # change the ending of the last added blink of blinksCombineFinal
                blinksCombineFinal[-1][1] = blinksCombine[ind+1][1]
                ind = ind + 1
            
            
        if len(blinksCombine) == 1:
            blinksCombineFinal = blinksCombine.copy()
            
        
        if blinksCombine[-1][1] != blinksCombineFinal[-1][1]:
            if blinksCombine[-1][0] == blinksCombineFinal[-1][1]:
                blinksCombineFinal[-1][1] = blinksCombine[-1][1]
            else:
                blinksCombineFinal.append(blinksCombine[-1])
        
        
        
        
    #    for w, z in groupby(sorted(list(blinksCombine)), lambda x, y=itertools.count(): next(y)-x):
    #        group = list(z)
    #        blinksCombineFinal.append(tuple((group[0], group[-1])))
        
        for x in sorted(blinksCombineFinal, reverse=True):
            new_start = blink_missingData_startEndTuple[x[0]][0] 
            new_end = blink_missingData_startEndTuple[x[1]][1] 
            
            x_start = x[0]
            x_end = x[1]
            
            # delete also the blinkStart_tupleList_wLength, since it is going to be used to compute other metrics
            for blinkRemove in range(x[1], x[0]-1, -1):
                del blink_missingData_startEndTuple[blinkRemove]
            
            blink_missingData_startEndTuple.insert(x[0], tuple((new_start, new_end)))
    
    
    #blinkAndNonBlinkDurationList = [length/90 for start, length in blinkStart_tupleList_wLength]
    #timeInS_Trial_filter = timeInS_Trial[-1] - sum(blinkAndNonBlinkDurationList) 
    
    
    # remove blinks from data
    for blinkStart, blinkEnd in blink_missingData_startEndTuple:
        pupilData_df.loc[blinkStart:blinkEnd,'pupilLeft'] = np.nan
        pupilData_df.loc[blinkStart:blinkEnd,'pupilRight'] = np.nan
        replaceTrueList = range(blinkStart, blinkEnd+1, 1)
        interpolatedNan_bool[replaceTrueList] = True
    
    
    
    pupilData_df['pupilLeft'] = pupilData_df['pupilLeft'].astype(float).interpolate('linear', limit_direction = 'both')
    pupilData_df['pupilRight'] = pupilData_df['pupilRight'].astype(float).interpolate('linear', limit_direction = 'both')
    
    if pupilData_df.isnull().any().any():
        print('nan values in filtered data')
        #for i,val in enumerate(pupilData_filter[0:5000]):
        #    print(i, val, pupilData_woSingleMissingData[i])
        
    
    return pupilData_df, interpolatedNan_bool, missingData_blinks_bool, missingDataOverall_bool, blinkStart_tupleList_wLength_maxBlinkLength, blinkStart_tupleList_wLength_allMissingData


# In[ ]:


def FilterPupilSize_wBlinkData(GazeLog, TimeTyping, subjectAndSessionName):
    # function that uses the list of start and end trial times to find the pupil sizes for those trials and plots them
    
    # first create a list of times in gaze log
    timeStrGazeLog = [item3[0] for item3 in GazeLog]
    # convert the list of strings to datetime formats
    timeGazeLog = timeConversion(timeStrGazeLog)
    
    # internal time, to depict seconds
    timeInternalGazeLog = [float(item3[1]) for item3 in GazeLog]
    
    # extract pupil sizes in decimals from the strange 2 columns for every pupil
    pupil_indWrtValidityL = [1, 2]
    pupil_validityL = 4
    pupilLogL_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityL, pupil_validityL)
    
    pupil_indWrtValidityR = [1, 2]
    pupil_validityR = 5
    pupilLogR_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityR, pupil_validityR)
    
    
    # reduce the data to start and end of typing time
    timeTyping_start, timeTyping_startInd = nearestTimePoint(timeGazeLog, TimeTyping['startTime'])
    timeTyping_end, timeTyping_endInd = nearestTimePoint(timeGazeLog, TimeTyping['endTime'])
    
    
    pupilLogL_wDefinedTime = pupilLogL_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    pupilLogR_wDefinedTime = pupilLogR_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    
    timeGazeLog_wDefinedTime = timeGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    
    timeInS_GazeLog_wDefinedTime = timeInternalGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    timeInS_Difference = [(t - s)/1000000 for s, t in zip(timeInS_GazeLog_wDefinedTime, timeInS_GazeLog_wDefinedTime[1:])]
    timeInS_Difference.insert(0, 0)
    
    
    #timeInS = [sum(timeInS_Difference[:i]) for i, v in enumerate(timeInS_Difference)]
    
    
    #pupilData_df, interpolated_items, missingData_fromBlinks, missingData, missingDataBlink_startNlength, \
    #missingDataAll_startNlength = filterBlinks_wBlinkData(pupilLogL_wDefinedTime, pupilLogR_wDefinedTime, timeGazeLog_wDefinedTime)
    
    pupilData_df, interpolated_items, missingData_fromBlinks, missingData, missingDataBlink_startNlength,     missingDataAll_startNlength = filterBlinks_wBlinkData(pupilLogL_raw['column1'], pupilLogR_raw['column1'], timeGazeLog)
    
    
    
    
    #timeGazeLog_plot = np.arange(0, timeInS[-1], 1/90)
    
    #plotPupilSize_checkFilter(pupilData_df, pupilLogL_wDefinedTime, blinkStartAndEnd, 'blink removal', subjectAndSessionName)
    
    
    
    
    pupilData_df_hampel = dict()
    
    pupilData_df_hampel = pupilData_df.copy()
    pupilData_df_hampel['pupilLeft'] = hampel(pupilData_df['pupilLeft'], 25, 3, False)
    pupilData_df_hampel['pupilRight'] = hampel(pupilData_df['pupilRight'], 25, 3, False)
        
    
        
    return pupilData_df_hampel, interpolated_items, missingData_fromBlinks, missingData, missingDataBlink_startNlength, missingDataAll_startNlength


# # Blink data

# In[ ]:


def GetBlinkPropertiesForEvents(EventTrial_index, PupilData_df, MissingData_startNlength):
    
    samplingFrequency = 90
    
    blinkFrequencyList = list()
    blinkDurationAverageList = list()
    blinkDurationTotalList = list()
    blinkCountList = list()
    interBlinkDurationList = list()
    
    # for every trial event, find the blinks during that event
    for ind, eventStart in enumerate(EventTrial_index['start']):
        #print('eventStart: ', eventStart, PupilData_df['timeStamp'][eventStart])
        blinkInd_missingData = np.array([indBlink for indBlink in range(0, len(MissingData_startNlength)) if                 MissingData_startNlength[indBlink][0] > eventStart and MissingData_startNlength[indBlink][0] <                                          EventTrial_index['end'][ind]])
        blinkDurations_trial = np.array([MissingData_startNlength[indBlink][1] for indBlink in blinkInd_missingData])
        interBlinkDuration_trial = [MissingData_startNlength[indBlink][0] - MissingData_startNlength[indBlink-1][0] -                                   MissingData_startNlength[indBlink-1][1] for indBlink in blinkInd_missingData[1:]]
        
        
        
        trialTime = ((PupilData_df['timeStamp'][EventTrial_index['end'][ind]] - PupilData_df['timeStamp'][                                EventTrial_index['start'][ind]]).total_seconds())/60
        
        blinkCountList.append(len(blinkInd_missingData))
        blinkFrequencyList.append(len(blinkInd_missingData)/trialTime)
        
        blinkDurationTotalList.append(sum(blinkDurations_trial)/samplingFrequency)
        
        interBlinkDurationList.append(np.mean(interBlinkDuration_trial)/samplingFrequency)
        
        blinkDurationAverageList.append(np.mean(blinkDurations_trial)/samplingFrequency)
        MissingData_startNlength_array = np.array(MissingData_startNlength)
        #print(MissingData_startNlength_array[blinkInd_missingData])
        #print(len(blinkInd_missingData)/trialTime)
        #print(sum(blinkDurations_trial)/samplingFrequency)
        #print(np.mean(blinkDurations_trial)/samplingFrequency)
        
    Blink_df = pd.DataFrame(list(zip(blinkCountList, blinkFrequencyList, blinkDurationAverageList, blinkDurationTotalList,                                     interBlinkDurationList)), columns=['blinkCount', 'blinkFrequency',                                     'blinkDurationAverage', 'blinkDurationTotal', 'interBlinkDuration'])
    Blink_df = Blink_df.fillna(0)
    
    
    return Blink_df


# # Blink burst functions

# In[ ]:


def filterBlinks_wBlinkDataAndBurst(pupilDataL, pupilDataR, timeInDatetime):
    # filter any blinks and nan values lasting around 250ms (on average)
    # first the single nan occurances are replaced with mean of the values on either sides, 
    # as they are assumed to be from hardware problems
    # for the rest of the blinks, 250ms before and after the nan values are interpolated with a linear function
    # returns a dataframe with pupil size, and timestamp
    # http://faculty.washington.edu/chudler/facts.html
   
    
    # create a dataframe from the pupilsize and time
    pupilData_df = pd.DataFrame(list(zip(timeInDatetime, pupilDataL, pupilDataR)), columns=['timeStamp', 'pupilLeft', 'pupilRight'])
    
    # blink is every nan value in the range of 100-400ms 
    # 250 ms (22 samples) before and after the blink will also be removed
    extraBlinkSamples = 18
    
    
    #pupilData_woSingleMissingData = pupilData.copy()
    #timeList_woSingleMissingData = timeInDatetime.copy()
    #timeInS_woSingleMissingData = timeInS_Trial[-1]
    
    # in case of single missing data, that are due to hardware error, replace with the mean of the pupil size before and
    # after the nan value
    # missing values will be the same for left and right pupil
    missingVal_Single = np.argwhere(np.isnan(pupilDataL))
    missingVal_Single = list(itertools.chain.from_iterable(missingVal_Single)) # flatten the list 
    
    
    
    # if no blinks present, return the data
    if len(missingVal_Single) == 0:
        interpolatedNan_bool = np.array([False]*len(pupilData_df['pupilLeft']))
        return pupilData_df, interpolatedNan_bool
    
    # find the index and values to replace for single nan values
    pupilData_tuples_replaceSingleNan_left = [(val, np.mean([pupilDataL[val-1], pupilDataL[val+1]])) for i, val in enumerate(missingVal_Single) if (val != 0 and val != (len(pupilDataL)-1)) if not np.isnan(pupilDataL[val-1]) and not np.isnan(pupilDataL[val+1])]
    pupilData_tuples_replaceSingleNan_right = [(val, np.mean([pupilDataR[val-1], pupilDataR[val+1]])) for i, val in enumerate(missingVal_Single) if (val != 0 and val != (len(pupilDataR)-1)) if not np.isnan(pupilDataR[val-1]) and not np.isnan(pupilDataR[val+1])]
    
    
    interpolatedNan_bool = np.array([True if ind in dict(pupilData_tuples_replaceSingleNan_left) else False for ind, val in enumerate(pupilDataL)])
    missingData_single = interpolatedNan_bool
    
    # replace the single nan values with the mean of the pupil size on either sides
    indList = -1
    for ind, val in pupilData_tuples_replaceSingleNan_left:
        indList = indList + 1
        pupilData_df.iloc[ind, pupilData_df.columns.get_loc('pupilLeft')] = val
        pupilData_df.iloc[ind, pupilData_df.columns.get_loc('pupilRight')] = pupilData_tuples_replaceSingleNan_right[indList][1]
        
    
    # again, find the nan values in the pupil size
    # the list missingVal_SingleDifference contains the index of the first blink, followed by the difference in the index
    # to the next nan value
    
    
    
    # find the nan values again from pupilData['pupilLeft']
    missingVal_Rest_trueFalse = pupilData_df['pupilLeft'].isnull()
    missingVal_Rest = [i for i, x in enumerate(missingVal_Rest_trueFalse) if x]
    
    # if no blinks left, return the current pupilData
    if len(missingVal_Rest) == 0:
        return pupilData_df, interpolatedNan_bool

    
    # in the blinks left, find when the blinks start by finding a difference in the consecutive values of the indices
    missingVal_RestDifference = [t - s for s, t in zip(missingVal_Rest, missingVal_Rest[1:])]
    missingVal_RestDifference.insert(0, missingVal_Rest[0])
    
    blinkStart_tupleList = [(ind, sum(missingVal_RestDifference[0:ind+1])) for ind, val in enumerate(missingVal_RestDifference) if val != 1]
    
    blinkStart_tupleList_wLength = list()
    
    # create a list of tuples of blink start index and the length of the blink
    ind = -1
    blinkLengthSum = 0
    for blink_ind, blinkStartInd in blinkStart_tupleList:
        ind = ind + 1
        if ind != len(blinkStart_tupleList) - 1:
            
            blinkLength = blinkStart_tupleList[ind+1][0]-blink_ind
            blinkLengthSum = blinkLengthSum + blinkLength
            
            blinkStart_tupleList_wLength.append(tuple((blinkStartInd, blinkLength)))
        else:
            # for the last blink -- all blink lengths summed and subtracted from the length of the list
            # missingVal_RestDifference 
            blinkLength = len(missingVal_RestDifference)-blinkLengthSum
            blinkStart_tupleList_wLength.append(tuple((blinkStartInd, blinkLength)))
     
    
    
    
    # create a vector with True if a blink was at the position
    samplingFrequency = 90
    blinkLengthMax = np.ceil(0.5*samplingFrequency)
    blinkLengthMin = np.ceil(0.075*samplingFrequency)
    blinkStart_tupleList_wLength_allMissingData = [(blinkStart, blinkLength) for blinkStart, blinkLength in blinkStart_tupleList_wLength]    
    blinkStart_tupleList_wLength_maxBlinkLength = [(blinkStart, blinkLength) for blinkStart, blinkLength in blinkStart_tupleList_wLength if blinkLength<blinkLengthMax and blinkLength>blinkLengthMin]
    
    missingData_blinks_bool = np.array([False]*len(pupilData_df['pupilLeft']))
    for blinkStart, blinkLength in blinkStart_tupleList_wLength_maxBlinkLength:
        blinkIndices = np.arange(blinkStart, blinkStart + blinkLength).astype(int)
        if blinkIndices[-1] > len(pupilDataL):
            blinkIndices = np.arange(blinkStart, len(pupilDataL)-1)
        missingData_blinks_bool[blinkIndices] = True
        
    
    
    missingData_blinks_boolAll = np.array([False]*len(pupilData_df['pupilLeft']))
    for blinkStart, blinkLength in blinkStart_tupleList_wLength:
        blinkIndices = np.arange(blinkStart, blinkStart + blinkLength).astype(int)
        if blinkIndices[-1] > len(pupilDataL):
            blinkIndices = np.arange(blinkStart, len(pupilDataL)-1)
        missingData_blinks_boolAll[blinkIndices] = True
        
        
    missingDataOverall_bool = missingData_single + missingData_blinks_boolAll
    
    # Check for blink bursts: A blink burst (BB) is defined as two or more blinks within 0.5–2.0 s [Horiuchi]
    blinkInd = 0
    blinkBurstList = list()
    blinkBurstEndTimeList = list()
    
    while blinkInd < len(blinkStart_tupleList_wLength_maxBlinkLength) - 2: # make sure we only go to the second last blink 
        blinkStart = blinkStart_tupleList_wLength_maxBlinkLength[blinkInd][0]
        
        #print(blinkInd, blinkStart)
        
        blinkBurst = list()
        blinkBurst.append(blinkStart)
        
        #print(blinkInd, blinkStart)
        
        for blinkNextInd in range(blinkInd+1, len(blinkStart_tupleList_wLength_maxBlinkLength)-1):
            blinkNextStart = blinkStart_tupleList_wLength_maxBlinkLength[blinkNextInd][0]
            blinkNextEnd = blinkStart_tupleList_wLength_maxBlinkLength[blinkNextInd][1] + blinkNextStart
            
            if blinkNextEnd - blinkStart < 2*90: #and blinkNextEnd - blinkStart >= 0.5*90: # defined in Horiuchi
                blinkBurst.append(blinkNextStart)
                #print('Added: ', blinkNextStart, blinkNextEnd, blinkNextEnd - blinkStart)
            #elif blinkNextStart - blinkStart_tupleList_wLength_maxBlinkLength[blinkNextInd][1] < 0.1*90:
            #    blinkBurst.append(blinkNextStart)
                #print('something with length', blinkNextStart)
            else:
                
                if len(blinkBurst) > 1:
                    #print('appending blink burst to list', blinkBurst)

                    blinkBurstList.append(blinkBurst)
                    #print('blink Bursts: ', len(blinkBurst), blinkBurst)
                    indBlinkBurst_end = blinkStart_tupleList_wLength_maxBlinkLength[blinkNextInd-1][0] +                     blinkStart_tupleList_wLength_maxBlinkLength[blinkNextInd-1][1]
                    blinkBurstEndTimeList.append(timeInDatetime[indBlinkBurst_end])
                break
            
        blinkInd = blinkNextInd
    
    blinkBurst_timeAndLength = dict()
    # convert the samples from blinkBurstList to time
    
    blinkBurst_timeAndLength['timeStampStart'] = [timeInDatetime[blinkBurst[0]] for blinkBurst in blinkBurstList]
    blinkBurst_timeAndLength['timeStampEnd'] = blinkBurstEndTimeList
    blinkBurst_timeAndLength['nBlinks'] = [len(blinkBurst) for blinkBurst in blinkBurstList]
        
    
    
    # add to vector with start and end of tuple
    #beforeAfterNan = [False]*len(pupilData_df['pupilLeft'])
    #for blinkStart, blinkLength in blinkStart_tupleList_wLength:
    #    beforeAfterNan[blinkStart] = True
    #    beforeAfterNan[blinkStart+blinkLength] = True
    #    #print('start and end points: ', pupilData_df['timeStamp'][blinkStart], pupilData_df['timeStamp'][blinkStart + blinkLength])
    
    
    # create lists with start and end values for the blinks, based on blinkStart_tupleList_wLength, regardless of the blink length
    blink_missingData_startList = [blinkStartInd - extraBlinkSamples if (blinkStartInd - extraBlinkSamples) > 0 else 0 for blinkStartInd, blinkLength in blinkStart_tupleList_wLength]
    blink_missingData_endList = [blinkStartInd + blinkLength + extraBlinkSamples if (blinkStartInd + blinkLength + extraBlinkSamples) < (len(pupilData_df['pupilLeft'])-1) else (len(pupilData_df['pupilLeft'])-1) for blinkStartInd, blinkLength in blinkStart_tupleList_wLength]
    # create a list of tuples from the start and end points of the blink
    blink_missingData_startEndTuple = [(blinkStart, blink_missingData_endList[ind]) for ind, blinkStart in enumerate(blink_missingData_startList)] 
    
    
    
        
    
    
    # check if blinks need to be combined - blinksCombine is a list of list of 2 elements, the index of the blinks that should be combined
    blinksCombine = [[ind, ind+1] for ind, blink in enumerate(blink_missingData_startEndTuple[0:-1]) if blink[1] > blink_missingData_startEndTuple[ind+1][0]]
        
    if blinksCombine:
        # combine blinks that need to be combined - if multiple consecutive blinks need to be removed: eg - [1, 2], [2, 3] 
        # are included in the blinksCombine, the combined version should be [1, 3] 
        blinksCombineFinal = list()
        ind = -1
        while ind < len(blinksCombine)-2:
            
            ind = ind + 1
            blinkCombining = blinksCombine[ind]
            blinksCombineFinal.append(blinkCombining)
            while ind < len(blinksCombine)-2 and blinkCombining[1] == blinksCombine[ind+1][0]:
                # change the ending of the last added blink of blinksCombineFinal
                blinksCombineFinal[-1][1] = blinksCombine[ind+1][1]
                ind = ind + 1
            
            
        if len(blinksCombine) == 1:
            blinksCombineFinal = blinksCombine.copy()
            
        
        if blinksCombine[-1][1] != blinksCombineFinal[-1][1]:
            if blinksCombine[-1][0] == blinksCombineFinal[-1][1]:
                blinksCombineFinal[-1][1] = blinksCombine[-1][1]
            else:
                blinksCombineFinal.append(blinksCombine[-1])
        
        
        
        
    #    for w, z in groupby(sorted(list(blinksCombine)), lambda x, y=itertools.count(): next(y)-x):
    #        group = list(z)
    #        blinksCombineFinal.append(tuple((group[0], group[-1])))
        
        for x in sorted(blinksCombineFinal, reverse=True):
            new_start = blink_missingData_startEndTuple[x[0]][0] 
            new_end = blink_missingData_startEndTuple[x[1]][1] 
            
            x_start = x[0]
            x_end = x[1]
            
            # delete also the blinkStart_tupleList_wLength, since it is going to be used to compute other metrics
            for blinkRemove in range(x[1], x[0]-1, -1):
                del blink_missingData_startEndTuple[blinkRemove]
            
            blink_missingData_startEndTuple.insert(x[0], tuple((new_start, new_end)))
    
    
    #blinkAndNonBlinkDurationList = [length/90 for start, length in blinkStart_tupleList_wLength]
    #timeInS_Trial_filter = timeInS_Trial[-1] - sum(blinkAndNonBlinkDurationList) 
    
    
    # remove blinks from data
    for blinkStart, blinkEnd in blink_missingData_startEndTuple:
        pupilData_df.loc[blinkStart:blinkEnd,'pupilLeft'] = np.nan
        pupilData_df.loc[blinkStart:blinkEnd,'pupilRight'] = np.nan
        replaceTrueList = range(blinkStart, blinkEnd+1, 1)
        interpolatedNan_bool[replaceTrueList] = True
    
    
    
    pupilData_df['pupilLeft'] = pupilData_df['pupilLeft'].astype(float).interpolate('linear', limit_direction = 'both')
    pupilData_df['pupilRight'] = pupilData_df['pupilRight'].astype(float).interpolate('linear', limit_direction = 'both')
    
    if pupilData_df.isnull().any().any():
        print('nan values in filtered data')
        #for i,val in enumerate(pupilData_filter[0:5000]):
        #    print(i, val, pupilData_woSingleMissingData[i])
        
    
    return pupilData_df, interpolatedNan_bool, missingData_blinks_bool, missingDataOverall_bool, blinkStart_tupleList_wLength_maxBlinkLength, blinkStart_tupleList_wLength_allMissingData, blinkBurst_timeAndLength


# In[ ]:


def FilterPupilSize_wBlinkDataAndBurst(GazeLog, TimeTyping, subjectAndSessionName):
    # function that uses the list of start and end trial times to find the pupil sizes for those trials and plots them
    
    # first create a list of times in gaze log
    timeStrGazeLog = [item3[0] for item3 in GazeLog]
    # convert the list of strings to datetime formats
    timeGazeLog = timeConversion(timeStrGazeLog)
    
    # internal time, to depict seconds
    timeInternalGazeLog = [float(item3[1]) for item3 in GazeLog]
    
    # extract pupil sizes in decimals from the strange 2 columns for every pupil
    pupil_indWrtValidityL = [1, 2]
    pupil_validityL = 4
    pupilLogL_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityL, pupil_validityL)
    
    pupil_indWrtValidityR = [1, 2]
    pupil_validityR = 5
    pupilLogR_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityR, pupil_validityR)
    
    
    # reduce the data to start and end of typing time
    timeTyping_start, timeTyping_startInd = nearestTimePoint(timeGazeLog, TimeTyping['startTime'])
    timeTyping_end, timeTyping_endInd = nearestTimePoint(timeGazeLog, TimeTyping['endTime'])
    
    
    pupilLogL_wDefinedTime = pupilLogL_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    pupilLogR_wDefinedTime = pupilLogR_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    
    timeGazeLog_wDefinedTime = timeGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    
    timeInS_GazeLog_wDefinedTime = timeInternalGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    timeInS_Difference = [(t - s)/1000000 for s, t in zip(timeInS_GazeLog_wDefinedTime, timeInS_GazeLog_wDefinedTime[1:])]
    timeInS_Difference.insert(0, 0)
    
    
    #timeInS = [sum(timeInS_Difference[:i]) for i, v in enumerate(timeInS_Difference)]
    
    
    #pupilData_df, interpolated_items, missingData_fromBlinks, missingData, missingDataBlink_startNlength, \
    #missingDataAll_startNlength = filterBlinks_wBlinkData(pupilLogL_wDefinedTime, pupilLogR_wDefinedTime, timeGazeLog_wDefinedTime)
    
    pupilData_df, interpolated_items, missingData_fromBlinks, missingData, missingDataBlink_startNlength,     missingDataAll_startNlength, blinkBursts = filterBlinks_wBlinkDataAndBurst(pupilLogL_raw['column1'], pupilLogR_raw['column1'], timeGazeLog)
    
    
    
    
    #timeGazeLog_plot = np.arange(0, timeInS[-1], 1/90)
    
    #plotPupilSize_checkFilter(pupilData_df, pupilLogL_wDefinedTime, blinkStartAndEnd, 'blink removal', subjectAndSessionName)
    
    
    
    
    pupilData_df_hampel = dict()
    
    pupilData_df_hampel = pupilData_df.copy()
    pupilData_df_hampel['pupilLeft'] = hampel(pupilData_df['pupilLeft'], 25, 3, False)
    pupilData_df_hampel['pupilRight'] = hampel(pupilData_df['pupilRight'], 25, 3, False)
        
    
        
    return pupilData_df_hampel, interpolated_items, missingData_fromBlinks, missingData, missingDataBlink_startNlength,missingDataAll_startNlength, blinkBursts


# # Long blink functions

# In[ ]:


def filterBlinks_wBlinkData_longBlinks(pupilDataL, pupilDataR, timeInDatetime):
    # filter any blinks and nan values lasting around 250ms (on average)
    # first the single nan occurances are replaced with mean of the values on either sides, 
    # as they are assumed to be from hardware problems
    # for the rest of the blinks, 250ms before and after the nan values are interpolated with a linear function
    # returns a dataframe with pupil size, and timestamp
    # http://faculty.washington.edu/chudler/facts.html
   
    
    # create a dataframe from the pupilsize and time
    pupilData_df = pd.DataFrame(list(zip(timeInDatetime, pupilDataL, pupilDataR)), columns=['timeStamp', 'pupilLeft', 'pupilRight'])
    
    # blink is every nan value in the range of 100-400ms 
    # 250 ms (22 samples) before and after the blink will also be removed
    extraBlinkSamples = 18
    
    
    #pupilData_woSingleMissingData = pupilData.copy()
    #timeList_woSingleMissingData = timeInDatetime.copy()
    #timeInS_woSingleMissingData = timeInS_Trial[-1]
    
    # in case of single missing data, that are due to hardware error, replace with the mean of the pupil size before and
    # after the nan value
    # missing values will be the same for left and right pupil
    missingVal_Single = np.argwhere(np.isnan(pupilDataL))
    missingVal_Single = list(itertools.chain.from_iterable(missingVal_Single)) # flatten the list 
    
    
    
    # if no blinks present, return the data
    if len(missingVal_Single) == 0:
        interpolatedNan_bool = np.array([False]*len(pupilData_df['pupilLeft']))
        return pupilData_df, interpolatedNan_bool
    
    # find the index and values to replace for single nan values
    pupilData_tuples_replaceSingleNan_left = [(val, np.mean([pupilDataL[val-1], pupilDataL[val+1]])) for i, val in enumerate(missingVal_Single) if (val != 0 and val != (len(pupilDataL)-1)) if not np.isnan(pupilDataL[val-1]) and not np.isnan(pupilDataL[val+1])]
    pupilData_tuples_replaceSingleNan_right = [(val, np.mean([pupilDataR[val-1], pupilDataR[val+1]])) for i, val in enumerate(missingVal_Single) if (val != 0 and val != (len(pupilDataR)-1)) if not np.isnan(pupilDataR[val-1]) and not np.isnan(pupilDataR[val+1])]
    
    
    interpolatedNan_bool = np.array([True if ind in dict(pupilData_tuples_replaceSingleNan_left) else False for ind, val in enumerate(pupilDataL)])
    missingData_single = interpolatedNan_bool
    
    # replace the single nan values with the mean of the pupil size on either sides
    indList = -1
    for ind, val in pupilData_tuples_replaceSingleNan_left:
        indList = indList + 1
        pupilData_df.iloc[ind, pupilData_df.columns.get_loc('pupilLeft')] = val
        pupilData_df.iloc[ind, pupilData_df.columns.get_loc('pupilRight')] = pupilData_tuples_replaceSingleNan_right[indList][1]
        
    
    # again, find the nan values in the pupil size
    # the list missingVal_SingleDifference contains the index of the first blink, followed by the difference in the index
    # to the next nan value
    
    
    
    # find the nan values again from pupilData['pupilLeft']
    missingVal_Rest_trueFalse = pupilData_df['pupilLeft'].isnull()
    missingVal_Rest = [i for i, x in enumerate(missingVal_Rest_trueFalse) if x]
    
    # if no blinks left, return the current pupilData
    if len(missingVal_Rest) == 0:
        return pupilData_df, interpolatedNan_bool

    
    # in the blinks left, find when the blinks start by finding a difference in the consecutive values of the indices
    missingVal_RestDifference = [t - s for s, t in zip(missingVal_Rest, missingVal_Rest[1:])]
    missingVal_RestDifference.insert(0, missingVal_Rest[0])
    
    blinkStart_tupleList = [(ind, sum(missingVal_RestDifference[0:ind+1])) for ind, val in enumerate(missingVal_RestDifference) if val != 1]
    
    blinkStart_tupleList_wLength = list()
    
    # create a list of tuples of blink start index and the length of the blink
    ind = -1
    blinkLengthSum = 0
    for blink_ind, blinkStartInd in blinkStart_tupleList:
        ind = ind + 1
        if ind != len(blinkStart_tupleList) - 1:
            
            blinkLength = blinkStart_tupleList[ind+1][0]-blink_ind
            blinkLengthSum = blinkLengthSum + blinkLength
            
            blinkStart_tupleList_wLength.append(tuple((blinkStartInd, blinkLength)))
        else:
            # for the last blink -- all blink lengths summed and subtracted from the length of the list
            # missingVal_RestDifference 
            blinkLength = len(missingVal_RestDifference)-blinkLengthSum
            blinkStart_tupleList_wLength.append(tuple((blinkStartInd, blinkLength)))
     
    
    
    
    # create a vector with True if a blink was at the position
    samplingFrequency = 90
    blinkLengthMax = np.ceil(0.5*samplingFrequency)
    blinkLengthMin = np.ceil(0.3*samplingFrequency)
    blinkStart_tupleList_wLength_allMissingData = [(blinkStart, blinkLength) for blinkStart, blinkLength in blinkStart_tupleList_wLength]    
    blinkStart_tupleList_wLength_maxBlinkLength = [(blinkStart, blinkLength) for blinkStart, blinkLength in blinkStart_tupleList_wLength if blinkLength<blinkLengthMax and blinkLength>blinkLengthMin]
    
    missingData_blinks_bool = np.array([False]*len(pupilData_df['pupilLeft']))
    for blinkStart, blinkLength in blinkStart_tupleList_wLength_maxBlinkLength:
        blinkIndices = np.arange(blinkStart, blinkStart + blinkLength).astype(int)
        if blinkIndices[-1] > len(pupilDataL):
            blinkIndices = np.arange(blinkStart, len(pupilDataL)-1)
        missingData_blinks_bool[blinkIndices] = True
        
    
    
    missingData_blinks_boolAll = np.array([False]*len(pupilData_df['pupilLeft']))
    for blinkStart, blinkLength in blinkStart_tupleList_wLength:
        blinkIndices = np.arange(blinkStart, blinkStart + blinkLength).astype(int)
        if blinkIndices[-1] > len(pupilDataL):
            blinkIndices = np.arange(blinkStart, len(pupilDataL)-1)
        missingData_blinks_boolAll[blinkIndices] = True
        
    missingDataOverall_bool = missingData_single + missingData_blinks_boolAll
    
    # add to vector with start and end of tuple
    #beforeAfterNan = [False]*len(pupilData_df['pupilLeft'])
    #for blinkStart, blinkLength in blinkStart_tupleList_wLength:
    #    beforeAfterNan[blinkStart] = True
    #    beforeAfterNan[blinkStart+blinkLength] = True
    #    #print('start and end points: ', pupilData_df['timeStamp'][blinkStart], pupilData_df['timeStamp'][blinkStart + blinkLength])
    
    
    # create lists with start and end values for the blinks, based on blinkStart_tupleList_wLength, regardless of the blink length
    blink_missingData_startList = [blinkStartInd - extraBlinkSamples if (blinkStartInd - extraBlinkSamples) > 0 else 0 for blinkStartInd, blinkLength in blinkStart_tupleList_wLength]
    blink_missingData_endList = [blinkStartInd + blinkLength + extraBlinkSamples if (blinkStartInd + blinkLength + extraBlinkSamples) < (len(pupilData_df['pupilLeft'])-1) else (len(pupilData_df['pupilLeft'])-1) for blinkStartInd, blinkLength in blinkStart_tupleList_wLength]
    # create a list of tuples from the start and end points of the blink
    blink_missingData_startEndTuple = [(blinkStart, blink_missingData_endList[ind]) for ind, blinkStart in enumerate(blink_missingData_startList)] 
    
    
    # check if blinks need to be combined - blinksCombine is a list of list of 2 elements, the index of the blinks that should be combined
    blinksCombine = [[ind, ind+1] for ind, blink in enumerate(blink_missingData_startEndTuple[0:-1]) if blink[1] > blink_missingData_startEndTuple[ind+1][0]]
        
    if blinksCombine:
        # combine blinks that need to be combined - if multiple consecutive blinks need to be removed: eg - [1, 2], [2, 3] 
        # are included in the blinksCombine, the combined version should be [1, 3] 
        blinksCombineFinal = list()
        ind = -1
        while ind < len(blinksCombine)-2:
            
            ind = ind + 1
            blinkCombining = blinksCombine[ind]
            blinksCombineFinal.append(blinkCombining)
            while ind < len(blinksCombine)-2 and blinkCombining[1] == blinksCombine[ind+1][0]:
                # change the ending of the last added blink of blinksCombineFinal
                blinksCombineFinal[-1][1] = blinksCombine[ind+1][1]
                ind = ind + 1
            
            
        if len(blinksCombine) == 1:
            blinksCombineFinal = blinksCombine.copy()
            
        
        if blinksCombine[-1][1] != blinksCombineFinal[-1][1]:
            if blinksCombine[-1][0] == blinksCombineFinal[-1][1]:
                blinksCombineFinal[-1][1] = blinksCombine[-1][1]
            else:
                blinksCombineFinal.append(blinksCombine[-1])
        
        
        
        
    #    for w, z in groupby(sorted(list(blinksCombine)), lambda x, y=itertools.count(): next(y)-x):
    #        group = list(z)
    #        blinksCombineFinal.append(tuple((group[0], group[-1])))
        
        for x in sorted(blinksCombineFinal, reverse=True):
            new_start = blink_missingData_startEndTuple[x[0]][0] 
            new_end = blink_missingData_startEndTuple[x[1]][1] 
            
            x_start = x[0]
            x_end = x[1]
            
            # delete also the blinkStart_tupleList_wLength, since it is going to be used to compute other metrics
            for blinkRemove in range(x[1], x[0]-1, -1):
                del blink_missingData_startEndTuple[blinkRemove]
            
            blink_missingData_startEndTuple.insert(x[0], tuple((new_start, new_end)))
    
    
    #blinkAndNonBlinkDurationList = [length/90 for start, length in blinkStart_tupleList_wLength]
    #timeInS_Trial_filter = timeInS_Trial[-1] - sum(blinkAndNonBlinkDurationList) 
    
    
    # remove blinks from data
    for blinkStart, blinkEnd in blink_missingData_startEndTuple:
        pupilData_df.loc[blinkStart:blinkEnd,'pupilLeft'] = np.nan
        pupilData_df.loc[blinkStart:blinkEnd,'pupilRight'] = np.nan
        replaceTrueList = range(blinkStart, blinkEnd+1, 1)
        interpolatedNan_bool[replaceTrueList] = True
    
    
    
    pupilData_df['pupilLeft'] = pupilData_df['pupilLeft'].astype(float).interpolate('linear', limit_direction = 'both')
    pupilData_df['pupilRight'] = pupilData_df['pupilRight'].astype(float).interpolate('linear', limit_direction = 'both')
    
    if pupilData_df.isnull().any().any():
        print('nan values in filtered data')
        #for i,val in enumerate(pupilData_filter[0:5000]):
        #    print(i, val, pupilData_woSingleMissingData[i])
        
    
    return pupilData_df, interpolatedNan_bool, missingData_blinks_bool, missingDataOverall_bool, blinkStart_tupleList_wLength_maxBlinkLength, blinkStart_tupleList_wLength_allMissingData


# In[ ]:


def FilterPupilSize_wBlinkDataLong(GazeLog, TimeTyping, subjectAndSessionName):
    # function that uses the list of start and end trial times to find the pupil sizes for those trials and plots them
    
    # first create a list of times in gaze log
    timeStrGazeLog = [item3[0] for item3 in GazeLog]
    # convert the list of strings to datetime formats
    timeGazeLog = timeConversion(timeStrGazeLog)
    
    # internal time, to depict seconds
    timeInternalGazeLog = [float(item3[1]) for item3 in GazeLog]
    
    # extract pupil sizes in decimals from the strange 2 columns for every pupil
    pupil_indWrtValidityL = [1, 2]
    pupil_validityL = 4
    pupilLogL_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityL, pupil_validityL)
    
    pupil_indWrtValidityR = [1, 2]
    pupil_validityR = 5
    pupilLogR_raw = gazeConvert2ColumnsTo1(GazeLog, pupil_indWrtValidityR, pupil_validityR)
    
    
    # reduce the data to start and end of typing time
    timeTyping_start, timeTyping_startInd = nearestTimePoint(timeGazeLog, TimeTyping['startTime'])
    timeTyping_end, timeTyping_endInd = nearestTimePoint(timeGazeLog, TimeTyping['endTime'])
    
    
    pupilLogL_wDefinedTime = pupilLogL_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    pupilLogR_wDefinedTime = pupilLogR_raw['column1'][timeTyping_startInd:timeTyping_endInd+1]
    
    timeGazeLog_wDefinedTime = timeGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    
    timeInS_GazeLog_wDefinedTime = timeInternalGazeLog[timeTyping_startInd:timeTyping_endInd+1]
    timeInS_Difference = [(t - s)/1000000 for s, t in zip(timeInS_GazeLog_wDefinedTime, timeInS_GazeLog_wDefinedTime[1:])]
    timeInS_Difference.insert(0, 0)
    
    
    #timeInS = [sum(timeInS_Difference[:i]) for i, v in enumerate(timeInS_Difference)]
    
    
    #pupilData_df, interpolated_items, missingData_fromBlinks, missingData, missingDataBlink_startNlength, \
    #missingDataAll_startNlength = filterBlinks_wBlinkData(pupilLogL_wDefinedTime, pupilLogR_wDefinedTime, timeGazeLog_wDefinedTime)
    
    pupilData_df, interpolated_items, missingData_fromBlinks, missingData, missingDataBlink_startNlength,     missingDataAll_startNlength = filterBlinks_wBlinkData_longBlinks(pupilLogL_raw['column1'], pupilLogR_raw['column1'], timeGazeLog)
    
    
    
    
    #timeGazeLog_plot = np.arange(0, timeInS[-1], 1/90)
    
    #plotPupilSize_checkFilter(pupilData_df, pupilLogL_wDefinedTime, blinkStartAndEnd, 'blink removal', subjectAndSessionName)
    
    
    
    
    pupilData_df_hampel = dict()
    
    pupilData_df_hampel = pupilData_df.copy()
    pupilData_df_hampel['pupilLeft'] = hampel(pupilData_df['pupilLeft'], 25, 3, False)
    pupilData_df_hampel['pupilRight'] = hampel(pupilData_df['pupilRight'], 25, 3, False)
        
    
        
    return pupilData_df_hampel, interpolated_items, missingData_fromBlinks, missingData, missingDataBlink_startNlength, missingDataAll_startNlength


# In[ ]:


def getGazePointInDegrees_wHeight(GazeLog):

    # Position of eyes in the UCS by Tobii (in mm)
    #print('gazeOrigin')
    gazeOriginUCS_validityL = 0
    gazeOriginUCS_indWrtValidityL = [1, 2, 3, 4, 5, 6]
    gazeOriginUCS_Left = gazeConvert2ColumnsTo1(GazeLog, gazeOriginUCS_indWrtValidityL, gazeOriginUCS_validityL)
    
    gazeOriginUCS_validityR = 1
    gazeOriginUCS_indWrtValidityR = [1, 2, 3, 4, 5, 6]
    gazeOriginUCS_Right = gazeConvert2ColumnsTo1(GazeLog, gazeOriginUCS_indWrtValidityR, gazeOriginUCS_validityR)
    
    # find the average of each to compute the gaze origin
    gazeOriginUCS = dict()
    gazeOriginUCS['x'] = np.array([(v+gazeOriginUCS_Right['column1'][i])/2 for i, v in enumerate(gazeOriginUCS_Left['column1'])])
    gazeOriginUCS['y'] = np.array([(v+gazeOriginUCS_Right['column2'][i])/2 for i, v in enumerate(gazeOriginUCS_Left['column2'])])
    gazeOriginUCS['z'] = np.array([(v+gazeOriginUCS_Right['column3'][i])/2 for i, v in enumerate(gazeOriginUCS_Left['column3'])])
    
    
    # Location of eye gaze on screen in the UCS by Tobii (in mm)
    #print('gazePoint')
    gazePointUCS_validityL = 2
    gazePointUCS_indWrtValidityL = [1, 2, 3, 4, 5, 6]
    gazePointUCS_Left = gazeConvert2ColumnsTo1(GazeLog, gazePointUCS_indWrtValidityL, gazePointUCS_validityL)
    
    gazePointUCS_validityR = 3
    gazePointUCS_indWrtValidityR = [1, 2, 3, 4, 5, 6]
    gazePointUCS_Right = gazeConvert2ColumnsTo1(GazeLog, gazePointUCS_indWrtValidityR, gazePointUCS_validityR)
    
    # find the average of gaze point in UCS
    gazePointUCS = dict()
    gazePointUCS['x'] = np.array([(v+gazePointUCS_Right['column1'][i])/2 for i, v in enumerate(gazePointUCS_Left['column1'])])
    gazePointUCS['y'] = np.array([(v+gazePointUCS_Right['column2'][i])/2 for i, v in enumerate(gazePointUCS_Left['column2'])])
    gazePointUCS['z'] = np.array([(v+gazePointUCS_Right['column3'][i])/2 for i, v in enumerate(gazePointUCS_Left['column3'])])
    
    # find the distance between the screen and eyes, dividing by 10 to get cm from mm 
    distanceEyeGaze = [(np.sqrt((gazePointUCS['x'][i]-gazeOriginUCS['x'][i])**2 + (gazePointUCS['y'][i]-gazeOriginUCS['y'][i])**2 + (gazePointUCS['z'][i]-gazeOriginUCS['z'][i])**2))/10 for i, v in enumerate(gazePointUCS['x'])]
    
    #plt.figure()
    #plt.plot(distanceEyeGaze)
    #plt.title('distance between eye and point of gaze')
    
    # Location of eye gaze on screen in the ADCS by Tobii (in arbitrary units)
    #print('gaze point on screen')
    gazePointADCS_indWrtValidityL = [-4, -3, -2, -1]
    gazePointADCS_validityL = 3
    gazePointADCS_Left_au = gazeConvert2ColumnsTo1(GazeLog, gazePointADCS_indWrtValidityL, gazePointADCS_validityL)
    
    
    gazePointADCS_indWrtValidityR = [-4, -3, -2, -1]
    gazePointADCS_validityR = 4
    gazePointADCS_Right_au = gazeConvert2ColumnsTo1(GazeLog, gazePointADCS_indWrtValidityR, gazePointADCS_validityR)
    
    
    # gazePointADCS is in arbitrary units and needs to be converted to cm  
    screenLength = 59
    screenWidth = 34.5
    
    gazePointADCS_Left_au['column1'] = [i*59 for i in gazePointADCS_Left_au['column1']]
    gazePointADCS_Left_au['column2'] = [i*34.5 for i in gazePointADCS_Left_au['column2']]
    
    gazePointADCS_Right_au['column1'] = [i*59 for i in gazePointADCS_Right_au['column1']]
    gazePointADCS_Right_au['column2'] = [i*34.5 for i in gazePointADCS_Right_au['column2']]
    
    gazePointADCS = dict()
    gazePointADCS['x'] = [(v+gazePointADCS_Right_au['column1'][i])/2 for i, v in enumerate(gazePointADCS_Left_au['column1'])]
    gazePointADCS['y'] = [(v+gazePointADCS_Right_au['column2'][i])/2 for i, v in enumerate(gazePointADCS_Left_au['column2'])]
    
    
    gazePoint_degrees = dict()
    gazePoint_degrees['x'] = np.array([v*screenLength/distanceEyeGaze[i] for i, v in enumerate(gazePointADCS['x'])])
    gazePoint_degrees['y'] = np.array([v*screenWidth/distanceEyeGaze[i] for i, v in enumerate(gazePointADCS['y'])])
    
    # first create a list of times in gaze log
    timeStrGazeLog = [item3[0] for item3 in GazeLog]
    # convert the list of strings to datetime formats
    timeGazeLog = np.array(timeConversion(timeStrGazeLog))
    
    gazePoint_degrees['timeStamp'] = timeGazeLog
    
    
    distanceVerticalEye = np.array(gazeOriginUCS['y'])/10
    
    return gazePoint_degrees, gazePointADCS, distanceVerticalEye


# In[ ]:





# ## Baseline Question data

# In[ ]:


def getBaselineQuestion(TimeTyping, KeysSelected_new):

    BaselineQuestion = dict()
    BaselineQuestion['start'], BaselineQuestion['end'] = list(), list()
    
    KeysSelected = [key[1] for key in KeysSelected_new]
    
    BaselineQuestionEndInd = KeysSelected.index('NextPhrase')
    
    time1, t1, t2 = KeysSelected_new[BaselineQuestionEndInd][0].partition('+')
    BaselineQuestionEndTime = datetime.datetime.strptime(re.sub('[:.T]','-',time1[:-1]), "%Y-%m-%d-%H-%M-%S-%f")
    
    print(TimeTyping['startTime'], BaselineQuestionEndTime)
    
    BaselineQuestion['start'] = TimeTyping['startTime']
    BaselineQuestion['end'] = BaselineQuestionEndTime
    
    return BaselineQuestion
            


# In[ ]:


def EventBaselineFromPupilData(EventTimeInKeys, PupilSize_df):
    
    EventTime, EventIndex = dict(), dict()
    
    EventTime['start'], EventIndex['start'] = nearestTimePoint(PupilSize_df['timeStamp'].tolist(),                                                                EventTimeInKeys['start'])
    EventTime['end'], EventIndex['end'] = nearestTimePoint(PupilSize_df['timeStamp'].tolist(), EventTimeInKeys['end'])
        
    return EventTime, EventIndex    


# In[ ]:


def FindBaseline_woNan(EndFound, trialEnd_startInd, PupilMean_df, trial_nanValues,                       TrialComputationInd, pupilSize_TrialEnd, overlappingInd):
    
    sampling_frequency = 90
    done = 0
    end = len(trial_nanValues)
    
    while not done and end >= TrialComputationInd:
        start = int(end - TrialComputationInd)
        NanValuesPresent = trial_nanValues[start:end]
        
        if sum(NanValuesPresent) < 1 and len(NanValuesPresent)>1:
                    
            TrialEnd_noNanStartInd = trialEnd_startInd + start
            TrialEnd_noNanEndInd = trialEnd_startInd + end
            done = 1
                    
            pupilSize_TrialEnd.append(np.mean(np.array(PupilMean_df['pupilMean_weighted'][TrialEnd_noNanStartInd:                            TrialEnd_noNanEndInd].tolist())))
                    
            #print('end', PupilMean_df['timeStamp'][TrialEnd_noNanStartInd], PupilMean_df['timeStamp'][\
            #                                                                                TrialEnd_noNanEndInd])
            EndFound = True
                    
            break
        
        #end = end - overlappingInd
        end = end - 1
        
    return EndFound, pupilSize_TrialEnd


# In[ ]:


def GetBaselineForEveryTrial(EventTrials_reading, PupilSize_df, NanValues, TimeBaseline, SubjectAndSessionName):

    subjectID = SubjectAndSessionName.split('__')[0]
    timeBaselineCheck = 5 # seconds
    sampling_frequency = 90 
    BaselineInd = TimeBaseline*sampling_frequency
    BufferTime = timeBaselineCheck/2
    BufferInd = int(BufferTime*sampling_frequency)
    pupilSize_baseline = list()
    
    for ind, startTimeInd in enumerate(EventTrials_reading['start']):
        
        # starting pupil size - start and end are reversed, because we come from the end
        endBaselineTime_NextPhraseEnds = PupilSize_df['timeStamp'][startTimeInd] 
        #print(endBaselineTime_NextPhraseEnds)
        
        
        baselineEnd_endInd = startTimeInd + int(TimeBaseline*sampling_frequency)
        baselineEnd_startInd = baselineEnd_endInd - int(timeBaselineCheck*sampling_frequency) - int(TimeBaseline*                                                                                        sampling_frequency)
        
        trial_nanValues = NanValues[baselineEnd_startInd:baselineEnd_endInd]
        #print(ind, PupilSize_df['timeStamp'][baselineEnd_startInd], PupilSize_df['timeStamp'][baselineEnd_endInd])
        pupilTime_trial = PupilSize_df['timeStamp'][baselineEnd_startInd:baselineEnd_endInd].tolist()
        
        
        overlapping_time = TimeBaseline/2 #s
        overlappingInd = int(overlapping_time*sampling_frequency)
        nIterations = int((timeBaselineCheck)/(TimeBaseline-overlapping_time))+1
        
        PupilSize_trial = PupilSize_df['timeStamp'][baselineEnd_startInd:baselineEnd_endInd].tolist()
        done = 0
        
        end = len(trial_nanValues)-1
        BaselineFound = False
        
        while BaselineFound == False:
            #print(ind, PupilSize_df['timeStamp'][baselineEnd_startInd], PupilSize_df['timeStamp'][baselineEnd_endInd])
            BaselineFound, pupilSize_baseline = FindBaseline_woNan(BaselineFound, baselineEnd_startInd, PupilSize_df,                     trial_nanValues, BaselineInd, pupilSize_baseline, overlappingInd)
            
            
            baselineEnd_endInd = baselineEnd_startInd + BufferInd
            baselineEnd_startInd = baselineEnd_endInd - int(timeBaselineCheck*sampling_frequency)
            trial_nanValues = NanValues[baselineEnd_startInd:baselineEnd_endInd]
            
      

    return pupilSize_baseline


# # Relative pupil size

# In[ ]:


def FindTrialStart_woNan(StartFound, trialStart_startInd, PupilMean_df, PupilBaseline, trial_nanValues,                          TrialComputationInd, pupilSize_TrialStart, overlappingTime):
    
    sampling_frequency = 90
    done = 0
    start = 0
    while not done and start + TrialComputationInd < len(trial_nanValues):
            
        end = start + TrialComputationInd
        
        NanValuesPresent = trial_nanValues[start:end]
        #print(PupilSize_df['timeStamp'][trialStart_startInd+start], PupilSize_df['timeStamp'][trialStart_startInd+end])
        
        if sum(NanValuesPresent) < 1 and len(NanValuesPresent)>1:
                   
            TrialStart_noNanStartInd = trialStart_startInd + start
            TrialStart_noNanEndInd = trialStart_startInd + end
            done = 1
                    
            pupilSize_TrialStart.append(np.mean(np.array(PupilMean_df['pupilMean_weighted'][TrialStart_noNanStartInd:                            TrialStart_noNanEndInd].tolist())-PupilBaseline))
                    
            #print('start', PupilMean_df['timeStamp'][TrialStart_noNanStartInd], PupilMean_df['timeStamp'][\
            #                                                                            TrialStart_noNanEndInd])
            StartFound = True
            break
            
        #start = start + int(overlappingTime*sampling_frequency)
        start = start + 1
        
        
    return StartFound, pupilSize_TrialStart


# In[ ]:


def FindTrialEnd_woNan(EndFound, trialEnd_startInd, PupilMean_df, PupilBaseline, trial_nanValues,                       TrialComputationInd, pupilSize_TrialEnd, overlapping_time):
    
    sampling_frequency = 90
    done = 0
    end = len(trial_nanValues)
    
    while not done and end >= TrialComputationInd:
        start = end - TrialComputationInd
        NanValuesPresent = trial_nanValues[start:end]
            
        if sum(NanValuesPresent) < 1 and len(NanValuesPresent)>1:
                    
            TrialEnd_noNanStartInd = trialEnd_startInd + start
            TrialEnd_noNanEndInd = trialEnd_startInd + end
            done = 1
                    
            pupilSize_TrialEnd.append(np.mean(np.array(PupilMean_df['pupilMean_weighted'][TrialEnd_noNanStartInd:                            TrialEnd_noNanEndInd].tolist())-PupilBaseline))
                    
            #print('end', PupilMean_df['timeStamp'][TrialEnd_noNanStartInd], PupilMean_df['timeStamp'][\
            #                                                                                TrialEnd_noNanEndInd])
            EndFound = True
                    
            break
        
        #end = end - int(overlapping_time*sampling_frequency)
        end = end - 1
        
    return EndFound, pupilSize_TrialEnd


# In[ ]:


def GetRelativePupilSize_startingAndEnding(EventTrialsInd, PupilMean_df, PupilBaselineList, TrialComputationTime,                                            NanValues):
    
    pupilSize_TrialStart = list()
    pupilSize_TrialEnd = list()
    
    sampling_frequency = 90
    TrialComputationCheckTime = 5
    TrialComputationCheckInd = int(TrialComputationCheckTime*sampling_frequency)
    TrialComputationInd = int(TrialComputationTime*sampling_frequency)
    
    BufferTime = 2.5
    BufferInd = int(BufferTime*sampling_frequency)
    
    for ind, trialStart_startInd in enumerate(EventTrialsInd['start']):
        
        # starting pupil size
        
        trialStart_endInd = trialStart_startInd + TrialComputationCheckInd
        trial_nanValues = NanValues[trialStart_startInd:trialStart_endInd]
        
        PupilRelative = np.array(PupilMean_df['pupilMean_weighted'][trialStart_startInd:                            EventTrialsInd['end'][ind]].tolist())-PupilBaselineList[ind]
        
        timeTrial = np.array(PupilMean_df['timeStamp'][trialStart_startInd:                            EventTrialsInd['end'][ind]].tolist())
        
        
        overlapping_time = TrialComputationTime/2 #s
        StartFound = False
        
        while StartFound == False:
            #print(StartFound)
            #print('trial: ', PupilMean_df['timeStamp'][trialStart_startInd], PupilMean_df['timeStamp'][\
            #                                                                        trialStart_endInd])

            StartFound, pupilSize_TrialStart = FindTrialStart_woNan(StartFound, trialStart_startInd, PupilMean_df,                PupilBaselineList[ind], trial_nanValues, TrialComputationInd, pupilSize_TrialStart, overlapping_time)
        
            #if StartFound == False:
            #    print('START NOT FOUND!!!')
            trialStart_startInd = trialStart_endInd - BufferInd # shifting the checking of nan values only by buffer 
            # time, not 5s TrialComputationCheckTime
            trialStart_endInd = trialStart_startInd + TrialComputationCheckInd 
            trial_nanValues = NanValues[trialStart_startInd:trialStart_endInd]
            
            
            
        
        
        # starting pupil size - start and end are reversed, because we come from the end
        trialEnd_endInd = EventTrialsInd['end'][ind]
        trialEnd_startInd = trialEnd_endInd - TrialComputationCheckInd
        trial_nanValues = NanValues[trialEnd_startInd:trialEnd_endInd]
        
        
        
        
        EndFound = False
        
        end = trialEnd_endInd
        done = 0
        while EndFound == False:
            #print('trial: ', PupilMean_df['timeStamp'][trialEnd_startInd], PupilMean_df['timeStamp'][\
            #                                                                        trialEnd_endInd])

            EndFound, pupilSize_TrialEnd = FindTrialEnd_woNan(EndFound, trialEnd_startInd, PupilMean_df,                     PupilBaselineList[ind], trial_nanValues, TrialComputationInd, pupilSize_TrialEnd, overlapping_time)
            
            
            #if EndFound == False:
            #    print('END NOT FOUND!!!')
            
            trialEnd_endInd = trialEnd_startInd + BufferInd
            trialEnd_startInd = trialEnd_endInd - TrialComputationCheckInd
            trial_nanValues = NanValues[trialEnd_startInd:trialEnd_endInd]
            
            
        
    return pupilSize_TrialStart, pupilSize_TrialEnd


# In[ ]:





# In[ ]:





# ## Performance data computation

# In[ ]:


def scratchPadPhraseEdit(phraseUser, subjName, full_path, picture):
    phraseUserEnd = list()
    
    
    
    for row_ind in range(0, len(phraseUser)):
        if row_ind!= 0 and phraseUser[row_ind][1] == '':
            if phraseUser[row_ind-1][1] != 'scratchPadText':
                if len(phraseUser[row_ind-1][1])>2: # this also removes any answers on the difficulty of the sentence
                    phraseUserEnd.append(phraseUser[row_ind-1])
                    #print(phraseUser[row_ind-1])
                    
    #print(' (1) phrases reduced to :       ', phraseUserEnd)
    # remove first two trials of baseline question and text composition
    if 'not_described' in picture or '2019-01-14-14-58-30' in full_path or '2019-02-06-12-37-45_2' in full_path or '2019-02-18-10-28-35_2' in full_path: 
        # yss has not copied one of the sentences, aq_session1_2 has not written the baseline question, ls2_session4_2 did not describe the picture
        del phraseUserEnd[0]
    else:
        del phraseUserEnd[0:2]
      
    session_folder_name = full_path.split('\\')[-1]
    
    if session_folder_name in dict_phraseUser:
        index_to_be_removed = dict_phraseUser[session_folder_name]
        #print('session in user phrases found')
    else:
        index_to_be_removed = []
        
    
    if index_to_be_removed:
        for index in sorted(index_to_be_removed, reverse=True):
            del phraseUserEnd[index]
        
    
    #print(' (2) phrases reduced to:       ', phraseUserEnd)
    
    # remove the initial rating of complexity, if they have written it:
    for index in range(0,len(phraseUserEnd)):
        sentence = phraseUserEnd[index][1]
        if sentence[0].isdigit():
            
            # if there is also a space after the digit:
            
            if sentence[1] == ' ':
                phraseUserEnd[index][1] = phraseUserEnd[index][1][2:]
            else:
                phraseUserEnd[index][1] = phraseUserEnd[index][1][1:]
            
        elif sentence[1].isdigit():
            if sentence[2] == ' ':
                phraseUserEnd[index][1] = phraseUserEnd[index][1][3:]
            else:
                phraseUserEnd[index][1] = phraseUserEnd[index][1][2:]
    
    #print(' (3) phrases reduced to:       ', phraseUserEnd)
    
    return phraseUserEnd


# In[ ]:


def stimPhrasesEdit(PhraseLog, full_path):
   
    # Now extract phrases from the phrase file
    phraseStim_Phrases = [item[1] for item in PhraseLog]
    
    # session name
    session_folder_name = full_path.split('\\')[-1]
        
    phraseStim_PhrasesReduced, phraseStim_timeReduced = zip(*[(x[0], PhraseLog[phraseStim_Phrases.index(x[0])][0]) for x in groupby(phraseStim_Phrases)])
    
    PhraseLogReduced = [[phraseStim_timeReduced[i], phraseStim_PhrasesReduced[i]] for i in range(0, len(phraseStim_PhrasesReduced))]
    
    
    if PhraseLogReduced[-1][1] == 'THE EXPERIMENT IS NOW DONE':
        del PhraseLogReduced[-1]
        
    if PhraseLogReduced[0][1] == 'phraseText':
        del PhraseLogReduced[0]

    # Here, we want only the sentences typed
    notSentencesToType = list()
    for index in range(0,len(PhraseLogReduced)):
        sentence = PhraseLogReduced[index][1]
        if 'Svar på følgende spørgsmål' in sentence or 'Answer the question:' in sentence or 'What is the complete name of your university?' in sentence or '(give a score between 1 and 7)' in sentence or sentence == '':
            notSentencesToType.append(index)    
    
    for index in sorted(notSentencesToType, reverse=True):
        del PhraseLogReduced[index]
    
    
    replacingList = []
    PhraseLogReduced = findAndRemoveTrials(session_folder_name, dict_phraseStim, PhraseLogReduced, replacingList)
    
    
    return PhraseLogReduced


# In[ ]:


# find minimum cost and the operations that give rise to it
def minValnInd(costOptions, flagSame):
    operator = list()
    unique_entries = set(costOptions)
    valInd = { value : [ i for i, v in enumerate(costOptions) if v == value ] for value in unique_entries }
    keyVal = list(valInd.keys())
    min_value = min(keyVal)
    
    if 0 in valInd[min_value]:
        operator.append('D')
    if 1 in valInd[min_value]:
        operator.append('I')
    if 2 in valInd[min_value]:
        if flagSame == 0:
            operator.append('S')
        else:
            operator.append('N')   
    flagSame = None    
    return min_value, ''.join(operator)


# In[ ]:


# Function to compute the MSD, with cost of 2 for substitution and 1 for insertion and deletion
costSub = 1
costIns = 1
costDel = 1

def levenshteinDist(phraseIn, phraseOut):
    
    lenStim = len(phraseIn)
    lenUser = len(phraseOut)
    costMatrix = np.zeros((lenStim+1, lenUser+1), dtype=int)
    MSDoperation = np.empty([lenStim+1, lenUser+1], dtype="U4")
    costMatrix[0,0:] = range(0, lenUser+1)
    costMatrix[0:,0] = range(0, lenStim+1)
    MSDoperation[0,0:] = 'I'
    MSDoperation[0:,0] = 'D'
    
    for i in range(1, len(phraseIn)+1):
        iP = i - 1
        for j in range(1, len(phraseOut)+1):
            jP = j - 1
            if phraseIn[iP].lower() == phraseOut[jP].lower():
                # Define the possible cost array
                costOptionArray = [costMatrix[i,j-1]+costDel, costMatrix[i-1,j]+costIns, costMatrix[i-1,j-1]] 
                flagSame = 1
            else:
                costOptionArray = [costMatrix[i,j-1]+costDel, costMatrix[i-1,j]+costIns, costMatrix[i-1,j-1]+costSub]
                flagSame = 0
            costMatrix[i,j], MSDoperation[i][j] = minValnInd(costOptionArray, flagSame)
    #print(costMatrix)
    return costMatrix[-1,-1]


# In[ ]:





# In[ ]:





# # Time computation

# In[ ]:


def EffectiveTimeFromUserKeys(UserKeys, Event, pathOfSession):
    
    UserKeysTimeStr = [val[0] for val in UserKeys]
    UserKeysTime = timeConversion(UserKeysTimeStr)
    
    timeActivation = 250
    
    # session name
    session_folder_name = pathOfSession.split('\\')[-1]
    
    timeTypingList = list()
    
    
    for indTrial, eventStart in enumerate(Event['start']):
        eventEnd = Event['end'][indTrial]
        eventStart_UserKeys, eventStartUserKeysInd = nearestTimePoint(UserKeysTime, eventStart)
        eventEnd_UserKeys, eventEnd_UserKeysInd = nearestTimePoint(UserKeysTime, eventEnd)
        time_trial = 0
        
        keySelectedPrevious = ''
        
        for indUserKey in range(eventStartUserKeysInd, eventEnd_UserKeysInd):
            
            if len(UserKeys[indUserKey][1]) == 1 or UserKeys[indUserKey][1] in list_keysToBeCounted or 'Suggestion' in UserKeys[indUserKey][1]:
                
                if len(UserKeys[indUserKey+1]) < 4 or float(UserKeys[indUserKey][3]) > float(UserKeys[indUserKey+1][3]) or UserKeys[indUserKey][2]==1:
                    
                    #print(UserKeys[indUserKey])
                    
                        
                    # Sometimes, in between selection of a letter, there is a small pause and activation of another letter,
                    # indicated by a '0' activation of the second letter during the first letter selection.
                    # This requires adding only the activation time for the second letter, not the dwell-time for the 
                    # current letter, which will be completed later on
                    
                    if len(UserKeys[indUserKey+1])>3 and len(UserKeys[indUserKey+2])>3:
                        if float(UserKeys[indUserKey][3]) > float(UserKeys[indUserKey+1][3]):
                            if float(UserKeys[indUserKey+1][3]) == 0:
                                if UserKeys[indUserKey+2][1] == UserKeys[indUserKey][1]:
                                    time_trial = time_trial + timeActivation
                                    #print('Only timeActivation added')
                                    if UserKeys[indUserKey][2]==1:
                                        keySelectedPrevious = UserKeys[indUserKey][1]
                                    continue
                    
                    if keySelectedPrevious == UserKeys[indUserKey][1]: # if the key selection is accidentally continued,
                        # activation time should not be counted
                        time_trial = time_trial + float(UserKeys[indUserKey][3])
                        #print('Current dwell-time added', keySelectedPrevious)
                        
                    else:
                        #print('timeActivation added')
                        time_trial = time_trial + float(UserKeys[indUserKey][3]) + timeActivation
                    
                    if UserKeys[indUserKey][2]==1:
                            keySelectedPrevious = UserKeys[indUserKey][1]
                            
        timeTypingList.append(time_trial/1000) # add the time in seconds
        
    
    return timeTypingList


# In[ ]:





# In[ ]:





# # Combine reading and writing

# In[ ]:


def CombineReadingWriting(EventReading, EventWriting, EventReading_index, EventWriting_index):
    
    EventTrial = copy.deepcopy(EventReading)
    EventTrialEnd = [endTime for endTime in EventWriting['end']]
    EventTrial['end'] = EventTrialEnd
    
    EventTrial_index = copy.deepcopy(EventReading_index)
    EventTrialEnd_index = [endTime for endTime in EventWriting_index['end']]
    EventTrial_index['end'] = EventTrialEnd_index
    
    
    return EventTrial, EventTrial_index


# # Eye movement data

# ## With distance and height of person

# In[ ]:


def getGazePointInDegrees_wDistanceFromScreen(GazeLog):

    # Position of eyes in the UCS by Tobii (in mm)
    #print('gazeOrigin')
    gazeOriginUCS_validityL = 0
    gazeOriginUCS_indWrtValidityL = [1, 2, 3, 4, 5, 6]
    gazeOriginUCS_Left = gazeConvert2ColumnsTo1(GazeLog, gazeOriginUCS_indWrtValidityL, gazeOriginUCS_validityL)
    
    gazeOriginUCS_validityR = 1
    gazeOriginUCS_indWrtValidityR = [1, 2, 3, 4, 5, 6]
    gazeOriginUCS_Right = gazeConvert2ColumnsTo1(GazeLog, gazeOriginUCS_indWrtValidityR, gazeOriginUCS_validityR)
    
    # find the average of each to compute the gaze origin
    gazeOriginUCS = dict()
    gazeOriginUCS['x'] = np.array([(v+gazeOriginUCS_Right['column1'][i])/2 for i, v in enumerate(gazeOriginUCS_Left['column1'])])
    gazeOriginUCS['y'] = np.array([(v+gazeOriginUCS_Right['column2'][i])/2 for i, v in enumerate(gazeOriginUCS_Left['column2'])])
    gazeOriginUCS['z'] = np.array([(v+gazeOriginUCS_Right['column3'][i])/2 for i, v in enumerate(gazeOriginUCS_Left['column3'])])
    
    distanceBetweenEyes = [np.sqrt((gazeOriginUCS_Left['column1'][i]-gazeOriginUCS_Right['column1'][i])**2 +                                    (gazeOriginUCS_Left['column2'][i]-gazeOriginUCS_Right['column2'][i])**2 +                                    (gazeOriginUCS_Left['column3'][i]-gazeOriginUCS_Right['column3'][i])**2)/10 for i, v in                          enumerate(gazeOriginUCS_Left['column1'])]
    
    
    
    # Location of eye gaze on screen in the UCS by Tobii (in mm)
    #print('gazePoint')
    gazePointUCS_validityL = 2
    gazePointUCS_indWrtValidityL = [1, 2, 3, 4, 5, 6]
    gazePointUCS_Left = gazeConvert2ColumnsTo1(GazeLog, gazePointUCS_indWrtValidityL, gazePointUCS_validityL)
    
    gazePointUCS_validityR = 3
    gazePointUCS_indWrtValidityR = [1, 2, 3, 4, 5, 6]
    gazePointUCS_Right = gazeConvert2ColumnsTo1(GazeLog, gazePointUCS_indWrtValidityR, gazePointUCS_validityR)
    
    # find the average of gaze point in UCS
    gazePointUCS = dict()
    gazePointUCS['x'] = np.array([(v+gazePointUCS_Right['column1'][i])/2 for i, v in enumerate(gazePointUCS_Left['column1'])])
    gazePointUCS['y'] = np.array([(v+gazePointUCS_Right['column2'][i])/2 for i, v in enumerate(gazePointUCS_Left['column2'])])
    gazePointUCS['z'] = np.array([(v+gazePointUCS_Right['column3'][i])/2 for i, v in enumerate(gazePointUCS_Left['column3'])])
    
    # find the distance between the screen and eyes, dividing by 10 to get cm from mm 
    distanceEyeGaze = [(np.sqrt((gazePointUCS['x'][i]-gazeOriginUCS['x'][i])**2 +                                 (gazePointUCS['y'][i]-gazeOriginUCS['y'][i])**2 +                                 (gazePointUCS['z'][i]-gazeOriginUCS['z'][i])**2))/10 for i, v in enumerate(                                gazePointUCS['x'])]
    
    #plt.figure()
    #plt.plot(distanceEyeGaze)
    #plt.title('distance between eye and point of gaze')
    
    # Location of eye gaze on screen in the ADCS by Tobii (in arbitrary units)
    #print('gaze point on screen')
    gazePointADCS_indWrtValidityL = [-4, -3, -2, -1]
    gazePointADCS_validityL = 3
    gazePointADCS_Left_au = gazeConvert2ColumnsTo1(GazeLog, gazePointADCS_indWrtValidityL, gazePointADCS_validityL)
    
    
    gazePointADCS_indWrtValidityR = [-4, -3, -2, -1]
    gazePointADCS_validityR = 4
    gazePointADCS_Right_au = gazeConvert2ColumnsTo1(GazeLog, gazePointADCS_indWrtValidityR, gazePointADCS_validityR)
    
    
    # gazePointADCS is in arbitrary units and needs to be converted to cm  
    screenLength = 59
    screenWidth = 34.5
    
    gazePointADCS_Left_au['column1'] = [i*59 for i in gazePointADCS_Left_au['column1']]
    gazePointADCS_Left_au['column2'] = [i*34.5 for i in gazePointADCS_Left_au['column2']]
    
    gazePointADCS_Right_au['column1'] = [i*59 for i in gazePointADCS_Right_au['column1']]
    gazePointADCS_Right_au['column2'] = [i*34.5 for i in gazePointADCS_Right_au['column2']]
    
    gazePointADCS = dict()
    gazePointADCS['x'] = [(v+gazePointADCS_Right_au['column1'][i])/2 for i, v in enumerate(gazePointADCS_Left_au['column1'])]
    gazePointADCS['y'] = [(v+gazePointADCS_Right_au['column2'][i])/2 for i, v in enumerate(gazePointADCS_Left_au['column2'])]
    
    
    gazePoint_degrees = dict()
    gazePoint_degrees['x'] = np.array([v*screenLength/distanceEyeGaze[i] for i, v in enumerate(gazePointADCS['x'])])
    gazePoint_degrees['y'] = np.array([v*screenWidth/distanceEyeGaze[i] for i, v in enumerate(gazePointADCS['y'])])
    
    # first create a list of times in gaze log
    timeStrGazeLog = [item3[0] for item3 in GazeLog]
    # convert the list of strings to datetime formats
    timeGazeLog = np.array(timeConversion(timeStrGazeLog))
    
    gazePoint_degrees['timeStamp'] = timeGazeLog
    
    distanceEyeAndGazePoint = dict()
    distanceEyeAndGazePoint['distance'] = distanceBetweenEyes
    distanceEyeAndGazePoint['timeStamp'] = timeGazeLog
    
    return gazePoint_degrees, gazePointADCS, distanceBetweenEyes


# # Metric/parameter for every trial

# In[ ]:


def DistanceInEachTrial(distanceFromScreenTotal, eventWriting_index, baselineDistance_mean):
    
    
    distanceFromScreenMean = [np.nanmean(np.array(distanceFromScreenTotal[trialStart:eventWriting_index['end'][i]]) -                                         baselineDistance_mean) for i,trialStart in enumerate(eventWriting_index['start'])]
    distanceFromScreenMedian = [np.nanmedian(np.array(distanceFromScreenTotal[trialStart:eventWriting_index['end'][i]]) -                                         baselineDistance_mean) for i,trialStart in enumerate(eventWriting_index['start'])]
    distanceFromScreenMin = [np.nanmin(np.array(distanceFromScreenTotal[trialStart:eventWriting_index['end'][i]]) -                                         baselineDistance_mean) for i,trialStart in enumerate(eventWriting_index['start'])]
    distanceFromScreenMax = [np.nanmax(np.array(distanceFromScreenTotal[trialStart:eventWriting_index['end'][i]]) -                                         baselineDistance_mean) for i,trialStart in enumerate(eventWriting_index['start'])]
    
    
    return distanceFromScreenMean, distanceFromScreenMedian, distanceFromScreenMin, distanceFromScreenMax


# In[ ]:


def MetricForEveryTrial(timeLingeringOnKeysList, eventTrialsInKeysSelected):
    
    timeList = [key[0] for key in timeLingeringOnKeysList]
    timeBetweenKeys = [key[1] for key in timeLingeringOnKeysList]
    
    
    timeBetweenKeysTrialMean = list()
    timeBetweenKeysTrialStd = list()
    
    for i, timeTrialStart in enumerate(eventTrialsInKeysSelected['start']):
        timeTrialEnd = eventTrialsInKeysSelected['end'][i]
        timeTrial = (timeTrialEnd-timeTrialStart).total_seconds()
        
        timeList0 = list()
        
        
        for ind, timeOccurence in enumerate(timeList):
            
            if timeOccurence < timeTrialEnd: 
                if timeOccurence > timeTrialStart:
                    timeList0.append(timeBetweenKeys[ind])
            else:
                break
        if len(timeList0) > 0:
            print(timeList0)
            print(timeTrial)
            timeBetweenKeysTrialMean.append(np.mean(timeList0)/timeTrial)
            #timeBetweenKeysTrialStd.append(np.std(timeList0))
            
        else:
            timeBetweenKeysTrialMean.append(0)
            #timeBetweenKeysTrialStd.append(np.nan)
            
        timeList0 = list()
        
    return timeBetweenKeysTrialMean #, timeBetweenKeysTrialStd


# ## Data saving functions

# In[ ]:


class DataForEveryTrial:
    subjectID = ''
    blockNumber = ''
    sessionNumber = ''
    variable = ''
    dataForTrial = ''
    resultPathName = ''
   
    
    def printInfo(self):
        dataFrame = pd.DataFrame(list(zip([self.subjectID]*len(self.dataForTrial), [self.blockNumber]*len(self.dataForTrial), [self.sessionNumber]*len(self.dataForTrial), range(0,len(self.dataForTrial)+1), self.dataForTrial)), columns=['subjectID', 'block', 'session', 'trial', self.variable])
        
        return dataFrame
    
    def AddToFile(self):
        
        dataFrame = pd.DataFrame(list(zip([self.subjectID]*len(self.dataForTrial), [self.blockNumber]*len(self.dataForTrial), [self.sessionNumber]*len(self.dataForTrial), range(0,len(self.dataForTrial)+1), self.dataForTrial)), columns=['subjectID', 'block', 'session', 'trial', self.variable])
        book = load_workbook(self.resultPathName)
        writer = pd.ExcelWriter(self.resultPathName, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        startrow = writer.sheets['Sheet1'].max_row
        dataFrame.to_excel(writer, startrow = startrow, index = False, header = False)
        
        writer.save()
    


# In[ ]:


class DataForEverySession:
    subjectID = ''
    blockNumber = ''
    sessionNumber = ''
    variable = ''
    dataForTrial = ''
    resultPathName = ''
   
    
    def printInfo(self):
        dataFrame = pd.DataFrame(list(zip([self.subjectID], [self.blockNumber], [self.sessionNumber], [self.dataForTrial])), columns=['subjectID', 'block', 'session', self.variable])
        
        return dataFrame
    
    def AddToFile(self):
        
        dataFrame = pd.DataFrame(list(zip([self.subjectID], [self.blockNumber], [self.sessionNumber], [self.dataForTrial])), columns=['subjectID', 'block', 'session', self.variable])
        
        book = load_workbook(self.resultPathName)
        writer = pd.ExcelWriter(self.resultPathName, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        startrow = writer.sheets['Sheet1'].max_row
        dataFrame.to_excel(writer, startrow = startrow, index = False, header = False)
        
        writer.save()


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# # Old incorrect functions

# In[ ]:


def GetBaseline(EventTrials_reading, PupilSize_df, NanValues, TimeBaseline):
    
    timeBaselineCheck = 5 # seconds
    
    pupilSize_baseline = dict()
    pupilSize_baseline['pupilLeft'] = list()
    pupilSize_baseline['pupilRight'] = list()
    
    for ind, startTimeInd in enumerate(EventTrials_reading['start']):
        startBaselineTime_NextPhraseEnds = PupilSize_df['timeStamp'][startTimeInd]
        StartBaselineTime, StartBaselineInd = nearestTimePoint(PupilSize_df['timeStamp'].tolist(),                 startBaselineTime_NextPhraseEnds-datetime.timedelta(seconds=TimeBaseline))
        startBaselineTime_w5s = startBaselineTime_NextPhraseEnds + datetime.timedelta(seconds=timeBaselineCheck)
        
        EndBaselineTime, EndBaselineInd = nearestTimePoint(PupilSize_df['timeStamp'].tolist(), startBaselineTime_w5s)
        
        
        
        trial_pupilSize = PupilSize_df['timeStamp'][StartBaselineInd:EndBaselineInd]
        trial_nanValues = NanValues[StartBaselineInd:EndBaselineInd]
        
        # find baseline pupil size, for TimeBaseline 
        overlapping_time = 0.1 #s
        sampling_frequency = 90
        nIterations = int((timeBaselineCheck+TimeBaseline)/overlapping_time)+1
        iteration = 0
        done = 0
        while iteration < nIterations:
            NanValuesPresent = trial_nanValues[int(iteration*(TimeBaseline-overlapping_time)*sampling_frequency):                                     int((iteration*(TimeBaseline-overlapping_time) + TimeBaseline)*sampling_frequency)]
            
            if sum(NanValuesPresent) < 1 and len(NanValuesPresent)>1:
                #print('done')
                baselineStartInd = StartBaselineInd + int(iteration*(TimeBaseline-overlapping_time)*sampling_frequency)
                baselineEndInd = StartBaselineInd + int((iteration*(TimeBaseline-overlapping_time) + TimeBaseline)*                                                        sampling_frequency)
                done = 1
                
                #print(np.mean(PupilSize_df['pupilLeft'][baselineStartInd:baselineEndInd]))
                #print(np.mean(PupilSize_df['pupilRight'][baselineStartInd:baselineEndInd]))
                
                pupilSize_baseline['pupilLeft'].append(np.mean(PupilSize_df['pupilLeft'][baselineStartInd:baselineEndInd]))
                pupilSize_baseline['pupilRight'].append(np.mean(PupilSize_df['pupilRight'][baselineStartInd:baselineEndInd]))
        
                
                #plt.figure()
                #plt.plot(PupilSize_df['timeStamp'][baselineStartInd:baselineEndInd], PupilSize_df['pupilLeft'][\
                #    baselineStartInd:baselineEndInd])
                #plt.plot(PupilSize_df['timeStamp'][baselineStartInd:baselineEndInd], PupilSize_df['pupilRight'][\
                #    baselineStartInd:baselineEndInd])
                
                break
            
            iteration = iteration + 1
            
        if done < 1:
            print('baseline not found')
            # add another 5s to the baseline search
            
            extraBaseLineEndTime, extraBaseLineEndInd = nearestTimePoint(PupilSize_df['timeStamp'].tolist(), EndBaselineTime                    +datetime.timedelta(seconds=timeBaselineCheck))
            
            trial_nanValues = NanValues[EndBaselineInd:extraBaseLineEndInd]
            
            nIterations = int((timeBaselineCheck+TimeBaseline)/overlapping_time)+1
            iteration = 0
            done = 0
            while iteration < nIterations:
                NanValuesPresent = trial_nanValues[int(iteration*(TimeBaseline-overlapping_time)*sampling_frequency):                                     int((iteration*(TimeBaseline-overlapping_time) + TimeBaseline)*sampling_frequency)]
            
                if sum(NanValuesPresent) < 1 and len(NanValuesPresent)>1:
                    #print('done')
                    baselineStartInd = EndBaselineInd + int(iteration*(TimeBaseline-overlapping_time)*sampling_frequency)
                    baselineEndInd = EndBaselineInd + int((iteration*(TimeBaseline-overlapping_time) + TimeBaseline)*                                                        sampling_frequency)
                    done = 1
                
                    #print(np.mean(PupilSize_df['pupilLeft'][baselineStartInd:baselineEndInd]))
                    #print(np.mean(PupilSize_df['pupilRight'][baselineStartInd:baselineEndInd]))
                
                    pupilSize_baseline['pupilLeft'].append(np.mean(PupilSize_df['pupilLeft'][baselineStartInd:baselineEndInd]))
                    pupilSize_baseline['pupilRight'].append(np.mean(PupilSize_df['pupilRight'][baselineStartInd:baselineEndInd]))
        
                
                    #plt.figure()
                    #plt.plot(PupilSize_df['timeStamp'][baselineStartInd:baselineEndInd], PupilSize_df['pupilLeft'][\
                    #    baselineStartInd:baselineEndInd])
                    #plt.plot(PupilSize_df['timeStamp'][baselineStartInd:baselineEndInd], PupilSize_df['pupilRight'][\
                    #    baselineStartInd:baselineEndInd])
                
                    break
            
                iteration = iteration + 1
            
            if done < 1:
                print('baseline not found -- again')
                
            
            
    return pupilSize_baseline

